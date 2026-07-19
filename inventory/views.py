from rest_framework import viewsets, permissions, status, filters, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q, Count, F as models_F
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from django.template.loader import render_to_string
import csv
import io
import openpyxl

try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False

from inventory.models import (
    ItemCategory, Unit, UnitConversion, Brand, InventoryItem,
    CustomFieldDefinition, InventoryLocationType, InventoryLocation,
    StockLedger, StockSummary, InventoryTransfer,
    InventoryTransferHistory,
    InventoryAdjustment, InventoryAdjustmentReason,
    InventoryAdjustmentHistory,
    InventoryReservation, InventoryReservationHistory,
    InventoryReservationReason,
    StockCountReason, InventoryStockCount, InventoryStockCountItem,
    InventoryStockCountHistory, InventoryStockCountAttachment,
    PurchaseOrder, PurchaseOrderItem, PurchaseOrderHistory,
    PurchaseOrderAttachment, PurchaseReceipt, PurchaseReceiptItem,
    InventoryGoodsReceipt, InventoryGoodsReceiptItem,
    InventoryGoodsReceiptHistory, InventoryGoodsReceiptAttachment,
    InventorySupplierInvoice, InventorySupplierInvoiceItem,
    InventorySupplierInvoiceHistory, InventorySupplierInvoiceAttachment,
    InventoryPurchaseReturn, InventoryPurchaseReturnItem,
    InventoryPurchaseReturnHistory, InventoryPurchaseReturnAttachment,
    InventorySupplierPayment, InventorySupplierPaymentAllocation,
    InventorySupplierPaymentHistory, InventorySupplierPaymentAttachment,
)
from inventory.serializers import (
    ItemCategorySerializer, ItemCategoryListSerializer, CategoryTreeSerializer,
    UnitSerializer, UnitListSerializer, UnitConversionSerializer,
    BrandSerializer, BrandListSerializer,
    InventoryItemSerializer, InventoryItemListSerializer,
    CustomFieldDefinitionSerializer, BulkImportSerializer,
    LocationTypeSerializer, LocationTypeListSerializer,
    LocationSerializer, LocationListSerializer, LocationTreeSerializer,
    StockLedgerSerializer, CreateLedgerEntrySerializer,
    StockAvailabilitySerializer, LowStockSerializer,
    TransferListSerializer, TransferDetailSerializer,
    CreateTransferSerializer, ReceiveTransferSerializer,
    TransferHistorySerializer,
    AdjustmentReasonSerializer, AdjustmentReasonListSerializer,
    AdjustmentListSerializer, AdjustmentDetailSerializer,
    CreateAdjustmentSerializer, AdjustmentHistorySerializer,
    ReservationReasonSerializer, ReservationReasonListSerializer,
    ReservationListSerializer, ReservationDetailSerializer,
    CreateReservationSerializer, FulfillReservationSerializer,
    ReservationHistorySerializer,
    StockCountReasonSerializer, StockCountReasonListSerializer,
    StockCountListSerializer, StockCountDetailSerializer,
    CreateStockCountSerializer, UpdateStockCountSerializer,
    SaveCountProgressSerializer,
    AssignCountersSerializer, StockCountHistorySerializer,
    StockCountDifferenceSerializer,
    PurchaseOrderListSerializer, PurchaseOrderDetailSerializer,
    CreatePurchaseOrderSerializer, UpdatePurchaseOrderSerializer,
    PurchaseOrderItemSerializer, PurchaseOrderAttachmentSerializer,
    PurchaseOrderHistorySerializer, ReceivePurchaseSerializer,
    PurchaseReceiptListSerializer, PurchaseReceiptDetailSerializer,
    PurchaseReceiptItemSerializer,
    GRNListSerializer, GRNDetailSerializer,
    CreateGRNSerializer, UpdateGRNSerializer,
    GRNItemSerializer, GRNAttachmentSerializer, GRNHistorySerializer,
    SupplierInvoiceListSerializer, SupplierInvoiceDetailSerializer,
    CreateSupplierInvoiceSerializer, UpdateSupplierInvoiceSerializer,
    SupplierInvoiceItemSerializer, SupplierInvoiceAttachmentSerializer,
    SupplierInvoiceHistorySerializer, RecordPaymentSerializer,
    PurchaseReturnListSerializer, PurchaseReturnDetailSerializer,
    CreatePurchaseReturnSerializer, UpdatePurchaseReturnSerializer,
    PurchaseReturnItemSerializer, PurchaseReturnAttachmentSerializer,
    PurchaseReturnHistorySerializer,
    SupplierPaymentListSerializer, SupplierPaymentDetailSerializer,
    CreateSupplierPaymentSerializer, UpdateSupplierPaymentSerializer,
    SupplierPaymentAllocationSerializer, SupplierPaymentAttachmentSerializer,
    SupplierPaymentHistorySerializer, AllocatePaymentSerializer,
)
from inventory.permissions import (
    CanViewItems, CanCreateItems, CanEditItems,
    CanDeleteItems, CanImportItems, CanExportItems,
    CanViewCategories, CanCreateCategories, CanEditCategories, CanDeleteCategories,
    CanViewUnits, CanCreateUnits, CanEditUnits, CanDeleteUnits,
    CanViewBrands, CanCreateBrands, CanEditBrands, CanDeleteBrands,
    CanViewLocationTypes, CanCreateLocationTypes, CanEditLocationTypes, CanDeleteLocationTypes,
    CanViewLocations, CanCreateLocations, CanEditLocations, CanDeleteLocations,
    CanViewStock, CanExportStock, CanViewStockSnapshot, CanViewStockValuation,
    CanViewTransfers, CanCreateTransfers, CanEditTransfers, CanDeleteTransfers,
    CanSubmitTransfer, CanApproveTransfer, CanReceiveTransfer, CanExportTransfers,
    CanViewAdjustments, CanCreateAdjustments, CanEditAdjustments, CanDeleteAdjustments,
    CanSubmitAdjustment, CanApproveAdjustment, CanApplyAdjustment, CanExportAdjustments,
    CanViewReservations, CanCreateReservations, CanEditReservations, CanDeleteReservations,
    CanActivateReservations, CanFulfillReservations, CanCancelReservations, CanExportReservations,
    CanViewStockCounts, CanCreateStockCounts, CanEditStockCounts,
    CanAssignStockCounts, CanCountStockItems, CanSubmitStockCounts,
    CanApproveStockCounts, CanCancelStockCounts, CanExportStockCounts,
    CanPrintStockCounts,
    CanViewPurchaseOrders, CanCreatePurchaseOrders, CanEditPurchaseOrders,
    CanDeletePurchaseOrders, CanSendPurchaseOrders, CanReceivePurchaseOrders,
    CanCancelPurchaseOrders, CanClosePurchaseOrders, CanExportPurchaseOrders,
    CanPrintPurchaseOrders,
    CanViewGRNs, CanCreateGRNs, CanEditGRNs, CanDeleteGRNs,
    CanSubmitGRN, CanApproveGRN, CanReceiveGRN, CanCancelGRN,
    CanExportGRNs, CanPrintGRNs,
    CanViewSupplierInvoices, CanCreateSupplierInvoices,
    CanEditSupplierInvoices, CanDeleteSupplierInvoices,
    CanSubmitSupplierInvoice, CanApproveSupplierInvoice,
    CanPostSupplierInvoice, CanRecordPaymentSupplierInvoice,
    CanCancelSupplierInvoice, CanVoidSupplierInvoice,
    CanExportSupplierInvoices, CanPrintSupplierInvoices,
    CanViewPurchaseReturns, CanCreatePurchaseReturns,
    CanEditPurchaseReturns, CanDeletePurchaseReturns,
    CanSubmitPurchaseReturn, CanApprovePurchaseReturn,
    CanReturnPurchaseReturn, CanCompletePurchaseReturn,
    CanCancelPurchaseReturn,     CanExportPurchaseReturns,
    CanViewSupplierPayments, CanCreateSupplierPayments,
    CanEditSupplierPayments, CanDeleteSupplierPayments,
    CanSubmitSupplierPayment, CanApproveSupplierPayment,
    CanPostSupplierPayment, CanAllocateSupplierPayment,
    CanCancelSupplierPayment, CanVoidSupplierPayment,
    CanExportSupplierPayments,
    CanViewDashboard, CanViewReports, CanExportReports, CanPrintReports,
)
from datetime import datetime
from django.utils import timezone
from inventory.services.stock_engine import (
    get_all_availability, get_item_availability, get_location_availability,
    get_low_stock_items, get_out_of_stock_items,
    get_valuation, get_snapshot, create_ledger_entry,
)
from inventory.services.transfer_service import (
    create_transfer, submit_transfer, approve_transfer,
    reject_transfer, dispatch_transfer, receive_transfer,
    cancel_transfer,
)
from inventory.services.adjustment_service import (
    create_adjustment, update_adjustment,
    submit_adjustment, approve_adjustment, reject_adjustment,
    apply_adjustment, cancel_adjustment,
)
from inventory.services.reservation_service import (
    create_reservation, update_reservation,
    activate_reservation, fulfill_reservation,
    cancel_reservation, expire_reservation,
    bulk_cancel_reservations,
)
from inventory.services.stock_count_service import (
    create_stock_count, update_stock_count,
    assign_counters, start_counting,
    save_count_progress, submit_stock_count,
    approve_stock_count, complete_stock_count,
    cancel_stock_count, get_difference_summary,
    reload_items_from_ledger,
)

from inventory.services.purchase_order_service import (
    create_purchase_order, update_purchase_order,
    send_purchase_order, receive_purchase_order_items,
    close_purchase_order, cancel_purchase_order,
)
from inventory.services.grn_service import (
    create_grn, update_grn, submit_grn, approve_grn,
    receive_grn, complete_grn, cancel_grn, get_history,
)
from inventory.services.supplier_invoice_service import (
    create_invoice, update_invoice,
    submit_invoice, approve_invoice, post_invoice,
    record_payment, cancel_invoice, void_invoice, get_history as get_invoice_history,
)
from inventory.feature_decorators import InventoryFeatureMixin


def _get_tenant_id(request):
    """Helper to get tenant_id from request user.
    Returns None for unauthenticated users — the permission layer
    will reject the request before this matters for views.
    """
    if not request.user.is_authenticated:
        return None
    return request.user.organization_id


def _get_permissions_for_action(action_map, default):
    """Helper to resolve permission classes based on action."""
    def get_permissions(view):
        for action, perm_class in action_map.items():
            if view.action == action:
                return [perm_class()]
        return [default()]
    return get_permissions


# ============================================================================
# CATEGORY VIEWSET
# ============================================================================

class ItemCategoryViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    required_feature = "categories"
    queryset = ItemCategory.objects.all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'status': ['exact'],
        'parent': ['exact', 'isnull'],
    }
    search_fields = ['category_code', 'category_name', 'description']
    ordering_fields = ['category_code', 'category_name', 'created_at', 'status']
    ordering = ['category_name']

    def get_serializer_class(self):
        if self.action == 'list':
            return ItemCategorySerializer
        elif self.action == 'tree':
            return CategoryTreeSerializer
        return ItemCategorySerializer

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        qs = ItemCategory.objects.filter(tenant_id=tenant_id).select_related('parent')

        show_archived = self.request.query_params.get('show_archived', 'false') == 'true'
        if not show_archived and self.action != 'tree':
            qs = qs.exclude(status='ARCHIVED')

        return qs

    def get_permissions(self):
        perms_map = {
            'create': [CanCreateCategories],
            'update': [CanEditCategories],
            'partial_update': [CanEditCategories],
            'destroy': [CanDeleteCategories],
        }
        for action, perms in perms_map.items():
            if self.action == action:
                return [p() for p in perms]
        return [CanViewCategories()]

    def perform_create(self, serializer):
        user = self.request.user
        serializer.save(
            tenant_id=_get_tenant_id(self.request),
            created_by=user,
            updated_by=user,
        )

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    def perform_destroy(self, instance):
        instance.status = 'ARCHIVED'
        instance.save()

    @action(detail=True, methods=['post'])
    def archive(self, request, pk=None):
        category = self.get_object()
        category.status = 'ARCHIVED'
        category.save()
        return Response({
            'success': True,
            'message': f"Category '{category.category_name}' archived."
        })

    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        category = self.get_object()
        category.status = 'ACTIVE'
        category.save()
        return Response({
            'success': True,
            'message': f"Category '{category.category_name}' restored."
        })

    @action(detail=False, methods=['get'])
    def tree(self, request):
        tenant_id = _get_tenant_id(request)
        categories = ItemCategory.objects.filter(
            tenant_id=tenant_id,
            parent__isnull=True,
        )
        show_archived = request.query_params.get('show_archived', 'false') == 'true'
        if not show_archived:
            categories = categories.exclude(status='ARCHIVED')

        serializer = CategoryTreeSerializer(
            categories, many=True,
            context={'request': request, 'filters': {'status': None}}
        )
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        tenant_id = _get_tenant_id(request)
        qs = ItemCategory.objects.filter(tenant_id=tenant_id)

        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        fmt = request.query_params.get('export_format', 'csv').lower()
        if fmt == 'csv':
            return self._export_csv(qs)
        else:
            return self._export_excel(qs)

    def _export_csv(self, qs):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Category Code', 'Category Name', 'Parent', 'Description', 'Status'])
        for cat in qs:
            writer.writerow([
                cat.category_code,
                cat.category_name,
                cat.parent.category_code if cat.parent else '',
                cat.description,
                cat.status,
            ])
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="categories.csv"'
        return response

    def _export_excel(self, qs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Categories'
        ws.append(['Category Code', 'Category Name', 'Parent', 'Description', 'Status'])
        for cat in qs:
            ws.append([
                cat.category_code,
                cat.category_name,
                cat.parent.category_code if cat.parent else '',
                cat.description,
                cat.status,
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="categories.xlsx"'
        return response

    @action(detail=False, methods=['post'], url_path='import')
    def import_bulk(self, request):
        file = request.FILES.get('file')
        if file:
            return self._import_from_file(file, request)

        serializer = BulkImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        return self._bulk_create(serializer.validated_data['items'], request)

    def _import_from_file(self, file, request):
        tenant_id = _get_tenant_id(request)
        created = []
        errors = []
        try:
            if file.name.endswith('.csv'):
                reader = csv.DictReader(io.StringIO(file.read().decode('utf-8-sig')))
                rows = list(reader)
            elif file.name.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(file, read_only=True)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows())]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                return Response({'error': 'Unsupported file format.'}, status=status.HTTP_400_BAD_REQUEST)

            for idx, row in enumerate(rows):
                code = str(row.get('Category Code', row.get('category_code', ''))).strip()
                name = str(row.get('Category Name', row.get('category_name', ''))).strip()
                if not code or not name:
                    errors.append({'index': idx, 'error': 'Category Code and Name are required.'})
                    continue

                if ItemCategory.objects.filter(tenant_id=tenant_id, category_code=code).exists():
                    errors.append({'index': idx, 'error': f"Code '{code}' already exists."})
                    continue

                parent_code = str(row.get('Parent', row.get('parent', ''))).strip()
                parent = None
                if parent_code:
                    parent = ItemCategory.objects.filter(
                        tenant_id=tenant_id, category_code=parent_code
                    ).first()

                cat = ItemCategory(
                    tenant_id=tenant_id,
                    category_code=code,
                    category_name=name,
                    parent=parent,
                    description=str(row.get('Description', row.get('description', ''))).strip(),
                    status='ACTIVE',
                )
                cat.save()
                created.append(ItemCategoryListSerializer(cat).data)

        except Exception as e:
            return Response({'error': f'Failed to parse file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'created_count': len(created),
            'error_count': len(errors),
            'created': created,
            'errors': errors,
        }, status=status.HTTP_201_CREATED)

    def _bulk_create(self, items_data, request):
        tenant_id = _get_tenant_id(request)
        created = []
        errors = []

        for idx, item in enumerate(items_data):
            code = item.get('category_code', '').strip()
            name = item.get('category_name', '').strip()
            if not code or not name:
                errors.append({'index': idx, 'error': 'Category Code and Name are required.'})
                continue

            if ItemCategory.objects.filter(tenant_id=tenant_id, category_code=code).exists():
                errors.append({'index': idx, 'error': f"Code '{code}' already exists."})
                continue

            cat = ItemCategory(
                tenant_id=tenant_id,
                category_code=code,
                category_name=name,
                description=item.get('description', ''),
                status='ACTIVE',
            )
            cat.save()
            created.append(ItemCategoryListSerializer(cat).data)

        return Response({
            'created_count': len(created),
            'error_count': len(errors),
            'created': created,
            'errors': errors,
        }, status=status.HTTP_201_CREATED)


# ============================================================================
# UNIT VIEWSET
# ============================================================================

class UnitViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    required_feature = "units"
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'status': ['exact'],
    }
    search_fields = ['unit_code', 'unit_name', 'symbol']
    ordering_fields = ['unit_code', 'unit_name', 'created_at', 'status']
    ordering = ['unit_name']

    def get_serializer_class(self):
        if self.action == 'list':
            return UnitSerializer
        return UnitSerializer

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        qs = Unit.objects.filter(tenant_id=tenant_id)

        show_archived = self.request.query_params.get('show_archived', 'false') == 'true'
        if not show_archived:
            qs = qs.exclude(status='ARCHIVED')

        return qs

    def get_permissions(self):
        perms_map = {
            'create': [CanCreateUnits],
            'update': [CanEditUnits],
            'partial_update': [CanEditUnits],
            'destroy': [CanDeleteUnits],
        }
        for action, perms in perms_map.items():
            if self.action == action:
                return [p() for p in perms]
        return [CanViewUnits()]

    def perform_create(self, serializer):
        user = self.request.user
        serializer.save(
            tenant_id=_get_tenant_id(self.request),
        )

    def perform_destroy(self, instance):
        if instance.items.filter(status='ACTIVE').exists():
            raise serializers.ValidationError(
                "Cannot archive unit used by active items."
            )
        instance.status = 'ARCHIVED'
        instance.save()

    @action(detail=True, methods=['post'])
    def archive(self, request, pk=None):
        unit = self.get_object()
        if unit.items.filter(status='ACTIVE').exists():
            return Response({
                'error': f"Cannot archive '{unit.unit_name}' — it is used by active items."
            }, status=status.HTTP_400_BAD_REQUEST)
        unit.status = 'ARCHIVED'
        unit.save()
        return Response({
            'success': True,
            'message': f"Unit '{unit.unit_name}' archived."
        })

    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        unit = self.get_object()
        unit.status = 'ACTIVE'
        unit.save()
        return Response({
            'success': True,
            'message': f"Unit '{unit.unit_name}' restored."
        })

    @action(detail=True, methods=['get', 'post', 'delete'], url_path='conversions')
    def conversions(self, request, pk=None):
        unit = self.get_object()

        if request.method == 'GET':
            qs = UnitConversion.objects.filter(
                Q(from_unit=unit) | Q(to_unit=unit)
            )
            serializer = UnitConversionSerializer(qs, many=True, context={'request': request})
            return Response(serializer.data)

        elif request.method == 'POST':
            data = request.data.copy()
            data['from_unit'] = str(unit.id)
            serializer = UnitConversionSerializer(
                data=data,
                context={'request': request},
            )
            if serializer.is_valid():
                serializer.save(tenant_id=_get_tenant_id(request))
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'DELETE':
            conv_id = request.query_params.get('conversion_id')
            if not conv_id:
                return Response(
                    {'error': 'conversion_id query parameter is required.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            try:
                conv = UnitConversion.objects.get(id=conv_id, from_unit=unit)
                conv.delete()
                return Response({'success': True, 'message': 'Conversion deleted.'})
            except UnitConversion.DoesNotExist:
                return Response(
                    {'error': 'Conversion not found.'},
                    status=status.HTTP_404_NOT_FOUND,
                )

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        tenant_id = _get_tenant_id(request)
        qs = Unit.objects.filter(tenant_id=tenant_id)

        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        fmt = request.query_params.get('export_format', 'csv').lower()
        if fmt == 'csv':
            return self._export_csv(qs)
        else:
            return self._export_excel(qs)

    def _export_csv(self, qs):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Unit Code', 'Unit Name', 'Symbol', 'Description', 'Status'])
        for unit in qs:
            writer.writerow([unit.unit_code, unit.unit_name, unit.symbol, unit.description, unit.status])
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="units.csv"'
        return response

    def _export_excel(self, qs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Units'
        ws.append(['Unit Code', 'Unit Name', 'Symbol', 'Description', 'Status'])
        for unit in qs:
            ws.append([unit.unit_code, unit.unit_name, unit.symbol, unit.description, unit.status])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="units.xlsx"'
        return response

    @action(detail=False, methods=['post'], url_path='import')
    def import_bulk(self, request):
        file = request.FILES.get('file')
        tenant_id = _get_tenant_id(request)
        created = []
        errors = []

        try:
            if file:
                if file.name.endswith('.csv'):
                    reader = csv.DictReader(io.StringIO(file.read().decode('utf-8-sig')))
                    rows = list(reader)
                elif file.name.endswith(('.xlsx', '.xls')):
                    wb = openpyxl.load_workbook(file, read_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in next(ws.iter_rows())]
                    rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers, row)))
                else:
                    return Response({'error': 'Unsupported file format.'}, status=status.HTTP_400_BAD_REQUEST)
            else:
                serializer = BulkImportSerializer(data=request.data)
                serializer.is_valid(raise_exception=True)
                rows = serializer.validated_data['items']

            for idx, row in enumerate(rows):
                code = str(row.get('Unit Code', row.get('unit_code', ''))).strip()
                name = str(row.get('Unit Name', row.get('unit_name', ''))).strip()
                if not code or not name:
                    errors.append({'index': idx, 'error': 'Unit Code and Name are required.'})
                    continue
                if Unit.objects.filter(tenant_id=tenant_id, unit_code=code).exists():
                    errors.append({'index': idx, 'error': f"Code '{code}' already exists."})
                    continue
                unit = Unit(
                    tenant_id=tenant_id,
                    unit_code=code,
                    unit_name=name,
                    symbol=str(row.get('Symbol', row.get('symbol', ''))).strip(),
                    description=str(row.get('Description', row.get('description', ''))).strip(),
                    status='ACTIVE',
                )
                unit.save()
                created.append(UnitListSerializer(unit).data)

        except Exception as e:
            return Response({'error': f'Failed to parse file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'created_count': len(created),
            'error_count': len(errors),
            'created': created,
            'errors': errors,
        }, status=status.HTTP_201_CREATED)


# ============================================================================
# BRAND VIEWSET
# ============================================================================

class BrandViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    required_feature = "brands"
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'status': ['exact'],
    }
    search_fields = ['brand_code', 'brand_name']
    ordering_fields = ['brand_code', 'brand_name', 'created_at', 'status']
    ordering = ['brand_name']

    def get_serializer_class(self):
        if self.action == 'list':
            return BrandSerializer
        return BrandSerializer

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        qs = Brand.objects.filter(tenant_id=tenant_id)

        show_archived = self.request.query_params.get('show_archived', 'false') == 'true'
        if not show_archived:
            qs = qs.exclude(status='ARCHIVED')

        return qs

    def get_permissions(self):
        perms_map = {
            'create': [CanCreateBrands],
            'update': [CanEditBrands],
            'partial_update': [CanEditBrands],
            'destroy': [CanDeleteBrands],
        }
        for action, perms in perms_map.items():
            if self.action == action:
                return [p() for p in perms]
        return [CanViewBrands()]

    def perform_create(self, serializer):
        user = self.request.user
        serializer.save(
            tenant_id=_get_tenant_id(self.request),
            created_by=user,
            updated_by=user,
        )

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    def perform_destroy(self, instance):
        instance.status = 'ARCHIVED'
        instance.save()

    @action(detail=True, methods=['post'])
    def archive(self, request, pk=None):
        brand = self.get_object()
        brand.status = 'ARCHIVED'
        brand.save()
        return Response({
            'success': True,
            'message': f"Brand '{brand.brand_name}' archived."
        })

    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        brand = self.get_object()
        brand.status = 'ACTIVE'
        brand.save()
        return Response({
            'success': True,
            'message': f"Brand '{brand.brand_name}' restored."
        })

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        tenant_id = _get_tenant_id(request)
        qs = Brand.objects.filter(tenant_id=tenant_id)

        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        fmt = request.query_params.get('export_format', 'csv').lower()
        if fmt == 'csv':
            return self._export_csv(qs)
        else:
            return self._export_excel(qs)

    def _export_csv(self, qs):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Brand Code', 'Brand Name', 'Description', 'Website', 'Status'])
        for brand in qs:
            writer.writerow([
                brand.brand_code, brand.brand_name, brand.description,
                brand.website, brand.status,
            ])
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="brands.csv"'
        return response

    def _export_excel(self, qs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Brands'
        ws.append(['Brand Code', 'Brand Name', 'Description', 'Website', 'Status'])
        for brand in qs:
            ws.append([
                brand.brand_code, brand.brand_name, brand.description,
                brand.website, brand.status,
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="brands.xlsx"'
        return response

    @action(detail=False, methods=['post'], url_path='import')
    def import_bulk(self, request):
        file = request.FILES.get('file')
        tenant_id = _get_tenant_id(request)
        created = []
        errors = []

        try:
            if file:
                if file.name.endswith('.csv'):
                    reader = csv.DictReader(io.StringIO(file.read().decode('utf-8-sig')))
                    rows = list(reader)
                elif file.name.endswith(('.xlsx', '.xls')):
                    wb = openpyxl.load_workbook(file, read_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in next(ws.iter_rows())]
                    rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers, row)))
                else:
                    return Response({'error': 'Unsupported file format.'}, status=status.HTTP_400_BAD_REQUEST)
            else:
                serializer = BulkImportSerializer(data=request.data)
                serializer.is_valid(raise_exception=True)
                rows = serializer.validated_data['items']

            for idx, row in enumerate(rows):
                code = str(row.get('Brand Code', row.get('brand_code', ''))).strip()
                name = str(row.get('Brand Name', row.get('brand_name', ''))).strip()
                if not code or not name:
                    errors.append({'index': idx, 'error': 'Brand Code and Name are required.'})
                    continue
                if Brand.objects.filter(tenant_id=tenant_id, brand_code=code).exists():
                    errors.append({'index': idx, 'error': f"Code '{code}' already exists."})
                    continue
                brand = Brand(
                    tenant_id=tenant_id,
                    brand_code=code,
                    brand_name=name,
                    description=str(row.get('Description', row.get('description', ''))).strip(),
                    website=str(row.get('Website', row.get('website', ''))).strip(),
                    status='ACTIVE',
                )
                brand.save()
                created.append(BrandListSerializer(brand).data)

        except Exception as e:
            return Response({'error': f'Failed to parse file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'created_count': len(created),
            'error_count': len(errors),
            'created': created,
            'errors': errors,
        }, status=status.HTTP_201_CREATED)


# ============================================================================
# INVENTORY ITEM VIEWSET
# ============================================================================

class InventoryItemViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    required_feature = "items"
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'status': ['exact'],
        'category': ['exact'],
        'brand': ['exact'],
    }
    search_fields = ['item_code', 'item_name', 'description', 'tags']
    ordering_fields = ['item_code', 'item_name', 'created_at', 'updated_at', 'status']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return InventoryItemListSerializer
        return InventoryItemSerializer

    def get_queryset(self):
        user = self.request.user
        tenant_id = user.organization_id
        qs = InventoryItem.objects.filter(tenant_id=tenant_id)

        qs = qs.select_related('category', 'unit', 'brand')

        show_archived = self.request.query_params.get('show_archived', 'false') == 'true'
        if not show_archived:
            qs = qs.exclude(status='ARCHIVED')

        return qs

    def get_permissions(self):
        if self.action == 'create':
            permission_classes = [CanCreateItems]
        elif self.action in ['update', 'partial_update']:
            permission_classes = [CanEditItems]
        elif self.action == 'destroy':
            permission_classes = [CanDeleteItems]
        elif self.action == 'import_bulk':
            permission_classes = [CanImportItems]
        elif self.action == 'export':
            permission_classes = [CanExportItems]
        else:
            permission_classes = [CanViewItems]
        return [p() for p in permission_classes]

    def perform_create(self, serializer):
        user = self.request.user
        tenant_id = user.organization_id
        serializer.save(tenant_id=tenant_id, created_by=user, updated_by=user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    def perform_destroy(self, instance):
        instance.status = 'ARCHIVED'
        instance.save()

    @action(detail=True, methods=['post'])
    def clone(self, request, pk=None):
        original = self.get_object()
        new_code = request.data.get(
            'item_code',
            f"{original.item_code}-CLONE"
        )
        new_name = request.data.get(
            'item_name',
            f"{original.item_name} - New Variant"
        )

        if InventoryItem.objects.filter(
            tenant_id=request.user.organization_id,
            item_code=new_code
        ).exists():
            return Response(
                {'error': f"Item code '{new_code}' already exists."},
                status=status.HTTP_400_BAD_REQUEST
            )

        cloned = InventoryItem.objects.create(
            tenant_id=request.user.organization_id,
            item_code=new_code,
            item_name=new_name,
            category=original.category,
            sub_category=original.sub_category,
            unit=original.unit,
            brand=original.brand,
            description=original.description,
            status='ACTIVE',
            custom_fields=original.custom_fields,
            tags=original.tags,
            cloned_from=original,
        )

        serializer = self.get_serializer(cloned)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def archive(self, request, pk=None):
        item = self.get_object()
        item.status = 'ARCHIVED'
        item.save()
        return Response({'success': True, 'message': f"Item '{item.item_name}' archived."})

    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        item = self.get_object()
        item.status = 'ACTIVE'
        item.save()
        return Response({'success': True, 'message': f"Item '{item.item_name}' restored."})

    @action(detail=False, methods=['post'], url_path='import')
    def import_bulk(self, request):
        file = request.FILES.get('file')
        if file:
            return self._import_from_file(file, request)

        serializer = BulkImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        items_data = serializer.validated_data['items']

        return self._bulk_create_items(items_data, request)

    def _import_from_file(self, file, request):
        tenant_id = request.user.organization_id
        created = []
        errors = []

        try:
            if file.name.endswith('.csv'):
                reader = csv.DictReader(io.StringIO(file.read().decode('utf-8-sig')))
                rows = list(reader)
            elif file.name.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(file, read_only=True)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows())]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                return Response(
                    {'error': 'Unsupported file format. Use CSV or Excel.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            for idx, row in enumerate(rows):
                item_code = str(row.get('Item Code', row.get('item_code', ''))).strip()
                item_name = str(row.get('Item Name', row.get('item_name', ''))).strip()
                category_name = str(row.get('Category', row.get('category', ''))).strip()
                unit_name = str(row.get('Unit', row.get('unit', ''))).strip()
                brand_name = str(row.get('Brand', row.get('brand', ''))).strip()
                description = str(row.get('Description', row.get('description', ''))).strip()

                if not item_code or not item_name:
                    errors.append({'index': idx, 'error': 'Item Code and Item Name are required.'})
                    continue

                if InventoryItem.objects.filter(
                    tenant_id=tenant_id,
                    item_code=item_code
                ).exists():
                    errors.append({'index': idx, 'error': f"Item code '{item_code}' already exists."})
                    continue

                category = None
                if category_name:
                    category = ItemCategory.objects.filter(
                        category_name__iexact=category_name
                    ).first()

                unit = None
                if unit_name:
                    unit = Unit.objects.filter(
                        Q(unit_name__iexact=unit_name) |
                        Q(symbol__iexact=unit_name)
                    ).first()

                brand = None
                if brand_name:
                    brand = Brand.objects.filter(brand_name__iexact=brand_name).first()

                item = InventoryItem(
                    tenant_id=tenant_id,
                    item_code=item_code,
                    item_name=item_name,
                    category=category,
                    unit=unit,
                    brand=brand,
                    description=description,
                    imported_from=file.name,
                    status='ACTIVE',
                )
                item.save()
                created.append(InventoryItemListSerializer(item).data)

        except Exception as e:
            return Response(
                {'error': f'Failed to parse file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response({
            'created_count': len(created),
            'error_count': len(errors),
            'created': created,
            'errors': errors,
        }, status=status.HTTP_201_CREATED)

    def _bulk_create_items(self, items_data, request):
        tenant_id = request.user.organization_id
        created = []
        errors = []

        for idx, item_data in enumerate(items_data):
            item_code = item_data.get('item_code', '').strip()
            item_name = item_data.get('item_name', '').strip()

            if InventoryItem.objects.filter(
                tenant_id=tenant_id,
                item_code=item_code
            ).exists():
                errors.append({
                    'index': idx,
                    'item_code': item_code,
                    'error': f"Item code '{item_code}' already exists."
                })
                continue

            category = None
            cat_name = item_data.get('category', '')
            if cat_name:
                category = ItemCategory.objects.filter(category_name__iexact=cat_name).first()

            unit = None
            unit_name = item_data.get('unit', '')
            if unit_name:
                unit = Unit.objects.filter(
                    Q(unit_name__iexact=unit_name) |
                    Q(symbol__iexact=unit_name)
                ).first()

            brand = None
            brand_name = item_data.get('brand', '')
            if brand_name:
                brand = Brand.objects.filter(brand_name__iexact=brand_name).first()

            item = InventoryItem(
                tenant_id=tenant_id,
                item_code=item_code,
                item_name=item_name,
                category=category,
                unit=unit,
                brand=brand,
                description=item_data.get('description', ''),
                status='ACTIVE',
            )
            item.save()
            created.append(InventoryItemListSerializer(item).data)

        return Response({
            'created_count': len(created),
            'error_count': len(errors),
            'created': created,
            'errors': errors,
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        tenant_id = request.user.organization_id
        items = InventoryItem.objects.filter(tenant_id=tenant_id).select_related(
            'category', 'unit', 'brand'
        )

        status_filter = request.query_params.get('status')
        if status_filter:
            items = items.filter(status=status_filter)

        category_filter = request.query_params.get('category')
        if category_filter:
            items = items.filter(category_id=category_filter)

        fmt = request.query_params.get('export_format', 'csv').lower()

        if fmt == 'csv':
            return self._export_csv(items)
        else:
            return self._export_excel(items)

    def _export_csv(self, items):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            'Item Code', 'Item Name', 'Category', 'Sub Category',
            'Unit', 'Brand', 'Description', 'Status', 'Tags'
        ])

        for item in items:
            writer.writerow([
                item.item_code,
                item.item_name,
                item.category.category_name if item.category else '',
                item.sub_category,
                item.unit.unit_name if item.unit else '',
                item.brand.brand_name if item.brand else '',
                item.description,
                item.status,
                ', '.join(item.tags) if item.tags else '',
            ])

        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv'
        )
        response['Content-Disposition'] = 'attachment; filename="inventory_items.csv"'
        return response

    def _export_excel(self, items):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Inventory Items'
        ws.append([
            'Item Code', 'Item Name', 'Category', 'Sub Category',
            'Unit', 'Brand', 'Description', 'Status', 'Tags'
        ])

        for item in items:
            ws.append([
                item.item_code,
                item.item_name,
                item.category.category_name if item.category else '',
                item.sub_category,
                item.unit.unit_name if item.unit else '',
                item.brand.brand_name if item.brand else '',
                item.description,
                item.status,
                ', '.join(item.tags) if item.tags else '',
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="inventory_items.xlsx"'
        return response


# ============================================================================
# CUSTOM FIELD DEFINITION VIEWSET
# ============================================================================

class CustomFieldDefinitionViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    required_feature = "items"
    queryset = CustomFieldDefinition.objects.filter(is_active=True)
    serializer_class = CustomFieldDefinitionSerializer
    permission_classes = [CanViewItems]
    filter_backends = [filters.SearchFilter]
    search_fields = ['field_name', 'field_label']

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            permission_classes = [CanCreateItems]
        else:
            permission_classes = [CanViewItems]
        return [p() for p in permission_classes]


# ============================================================================
# STOCK LEDGER VIEWSET
# ============================================================================

class StockLedgerViewSet(InventoryFeatureMixin, viewsets.ReadOnlyModelViewSet):
    required_feature = "stock"
    """View stock ledger entries."""
    serializer_class = StockLedgerSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'item': ['exact'],
        'location': ['exact'],
        'transaction_type': ['exact'],
        'created_at': ['gte', 'lte', 'date'],
    }
    search_fields = ['item__item_code', 'item__item_name', 'description', 'reference_id']
    ordering_fields = ['created_at', 'transaction_type', 'quantity']
    ordering = ['-created_at']

    def get_queryset(self):
        return StockLedger.objects.filter(
            tenant_id=_get_tenant_id(self.request)
        ).select_related('item', 'location')

    def get_permissions(self):
        return [CanViewStock()]


# ============================================================================
# STOCK LEDGER ENTRY CREATE (POST only)
# ============================================================================

class StockLedgerEntryView(InventoryFeatureMixin, viewsets.ViewSet):
    required_feature = "stock"
    """Create stock ledger entries."""
    permission_classes = [CanCreateItems]

    def create(self, request):
        serializer = CreateLedgerEntrySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data
        entry = create_ledger_entry(
            tenant_id=_get_tenant_id(request),
            item_id=data['item_id'],
            transaction_type=data['transaction_type'],
            quantity=data['quantity'],
            location_id=data.get('location_id'),
            unit_cost=data.get('unit_cost'),
            reference_type=data.get('reference_type', ''),
            reference_id=data.get('reference_id', ''),
            description=data.get('description', ''),
            created_by=request.user,
        )

        return Response(
            StockLedgerSerializer(entry, context={'request': request}).data,
            status=status.HTTP_201_CREATED,
        )


# ============================================================================
# STOCK AVAILABILITY VIEW
# ============================================================================

class StockAvailabilityViewSet(InventoryFeatureMixin, viewsets.ViewSet):
    """
    required_feature = "stock"
    View stock availability — calculated from the ledger in real-time.
    Never reads stored quantities.
    """
    permission_classes = [CanViewStock]

    def get_permissions(self):
        if self.action in ['export']:
            return [CanExportStock()]
        elif self.action in ['snapshot']:
            return [CanViewStockSnapshot()]
        elif self.action in ['valuation']:
            return [CanViewStockValuation()]
        return [CanViewStock()]

    def list(self, request):
        tenant_id = _get_tenant_id(request)
        filters = {
            'search': request.query_params.get('search', ''),
            'category': request.query_params.get('category'),
            'brand': request.query_params.get('brand'),
            'status': request.query_params.get('status'),
        }
        filters = {k: v for k, v in filters.items() if v}

        results = get_all_availability(tenant_id, filters)

        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 50))
        start = (page - 1) * page_size
        end = start + page_size
        page_data = results[start:end]

        return Response({
            'count': len(results),
            'results': page_data,
        })

    def retrieve(self, request, pk=None):
        tenant_id = _get_tenant_id(request)
        data = get_item_availability(pk, tenant_id)
        return Response(data)

    @action(detail=False, methods=['get'], url_path='location/(?P<location_id>[^/.]+)')
    def by_location(self, request, location_id=None):
        tenant_id = _get_tenant_id(request)
        data = get_location_availability(location_id, tenant_id)
        return Response(data)

    @action(detail=False, methods=['get'], url_path='low-stock')
    def low_stock(self, request):
        tenant_id = _get_tenant_id(request)
        data = get_low_stock_items(tenant_id)
        return Response({'count': len(data), 'results': data})

    @action(detail=False, methods=['get'], url_path='out-of-stock')
    def out_of_stock(self, request):
        tenant_id = _get_tenant_id(request)
        data = get_out_of_stock_items(tenant_id)
        return Response({'count': len(data), 'results': data})

    @action(detail=False, methods=['get'], url_path='valuation')
    def valuation(self, request):
        tenant_id = _get_tenant_id(request)
        filters = {
            'category': request.query_params.get('category'),
            'brand': request.query_params.get('brand'),
        }
        filters = {k: v for k, v in filters.items() if v}
        data = get_valuation(tenant_id, filters)
        return Response(data)

    @action(detail=False, methods=['get'], url_path='snapshot')
    def snapshot(self, request):
        tenant_id = _get_tenant_id(request)
        as_of = request.query_params.get('as_of')
        if as_of:
            try:
                as_of_date = datetime.strptime(as_of, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {'error': 'Invalid date format. Use YYYY-MM-DD.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            as_of_date = timezone.now().date()

        data = get_snapshot(tenant_id, as_of_date)
        return Response(data)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        tenant_id = _get_tenant_id(request)
        results = get_all_availability(tenant_id)
        fmt = request.query_params.get('export_format', 'csv').lower()

        if fmt == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'Item Code', 'Item Name', 'Category', 'Brand',
                'Physical', 'Reserved', 'Available',
                'In Transit', 'Damaged',
                'Cost Price', 'Selling Price',
                'Cost Value', 'Selling Value',
            ])
            for r in results:
                writer.writerow([
                    r['item_code'], r['item_name'],
                    r['category_name'], r['brand_name'],
                    r['physical'], r['reserved'], r['available'],
                    r['in_transit'], r['damaged'],
                    r['cost_price'], r['selling_price'],
                    r['cost_value'], r['selling_value'],
                ])
            output.seek(0)
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="stock_availability.csv"'
            return response
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Stock Availability'
            ws.append([
                'Item Code', 'Item Name', 'Category', 'Brand',
                'Physical', 'Reserved', 'Available',
                'In Transit', 'Damaged',
                'Cost Price', 'Selling Price',
                'Cost Value', 'Selling Value',
            ])
            for r in results:
                ws.append([
                    r['item_code'], r['item_name'],
                    r['category_name'], r['brand_name'],
                    float(r['physical']), float(r['reserved']),
                    float(r['available']), float(r['in_transit']),
                    float(r['damaged']),
                    float(r['cost_price']) if r['cost_price'] else '',
                    float(r['selling_price']) if r['selling_price'] else '',
                    float(r['cost_value']), float(r['selling_value']),
                ])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="stock_availability.xlsx"'
            return response


# ============================================================================
# LOCATION TYPE VIEWSET
# ============================================================================

class LocationTypeViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    required_feature = "location-types"
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'status': ['exact'],
    }
    search_fields = ['type_code', 'type_name']
    ordering_fields = ['type_code', 'type_name', 'created_at', 'status']
    ordering = ['type_name']

    def get_serializer_class(self):
        if self.action == 'list':
            return LocationTypeSerializer
        return LocationTypeSerializer

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        qs = InventoryLocationType.objects.filter(tenant_id=tenant_id)
        return qs

    def get_permissions(self):
        perms_map = {
            'create': [CanCreateLocationTypes],
            'update': [CanEditLocationTypes],
            'partial_update': [CanEditLocationTypes],
            'destroy': [CanDeleteLocationTypes],
        }
        for action, perms in perms_map.items():
            if self.action == action:
                return [p() for p in perms]
        return [CanViewLocationTypes()]

    def perform_create(self, serializer):
        serializer.save(tenant_id=_get_tenant_id(self.request))

    def perform_destroy(self, instance):
        if instance.locations.exists():
            raise serializers.ValidationError(
                "Cannot delete location type in use by existing locations."
            )
        instance.delete()


# ============================================================================
# TRANSFER VIEWSET
# ============================================================================

class TransferViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    required_feature = "transfers"
    """Stock transfers with full workflow: submit → approve → dispatch → receive"""
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'status': ['exact'],
        'source_location': ['exact'],
        'destination_location': ['exact'],
        'transfer_type': ['exact'],
        'transfer_date': ['gte', 'lte'],
    }
    search_fields = ['transfer_number', 'remarks']
    ordering_fields = ['transfer_number', 'transfer_date', 'created_at', 'status']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return TransferListSerializer
        elif self.action in ['create']:
            return CreateTransferSerializer
        elif self.action in ['receive']:
            return ReceiveTransferSerializer
        return TransferDetailSerializer

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        return InventoryTransfer.objects.filter(tenant_id=tenant_id).select_related(
            'source_location', 'destination_location', 'created_by'
        ).prefetch_related('items__item')

    def get_permissions(self):
        perms_map = {
            'create': [CanCreateTransfers],
            'update': [CanEditTransfers],
            'partial_update': [CanEditTransfers],
            'destroy': [CanDeleteTransfers],
            'submit': [CanSubmitTransfer],
            'approve': [CanApproveTransfer],
            'reject': [CanApproveTransfer],
            'perform_dispatch': [CanApproveTransfer],
            'receive': [CanReceiveTransfer],
            'export': [CanExportTransfers],
        }
        for action, perms in perms_map.items():
            if self.action == action:
                return [p() for p in perms]
        return [CanViewTransfers()]

    def update(self, request, *args, **kwargs):
        transfer = self.get_object()
        if transfer.status in ('COMPLETED', 'CANCELLED', 'RECEIVED', 'PARTIALLY_RECEIVED', 'IN_TRANSIT', 'APPROVED'):
            return Response(
                {'error': f"Cannot edit a transfer in status '{transfer.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        transfer = self.get_object()
        if transfer.status != 'DRAFT':
            return Response(
                {'error': f"Cannot delete a transfer in status '{transfer.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().destroy(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        tenant_id = _get_tenant_id(request)
        try:
            transfer = create_transfer(tenant_id, serializer.validated_data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output_serializer = TransferDetailSerializer(
            transfer, context={'request': request}
        )
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        transfer = self.get_object()
        try:
            transfer = submit_transfer(transfer, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = TransferDetailSerializer(transfer, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        transfer = self.get_object()
        notes = request.data.get('notes', '')
        try:
            transfer = approve_transfer(transfer, request.user, notes)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = TransferDetailSerializer(transfer, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        transfer = self.get_object()
        notes = request.data.get('notes', '')
        try:
            transfer = reject_transfer(transfer, request.user, notes)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = TransferDetailSerializer(transfer, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='dispatch')
    def perform_dispatch(self, request, pk=None):
        transfer = self.get_object()
        try:
            transfer = dispatch_transfer(transfer, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = TransferDetailSerializer(transfer, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def receive(self, request, pk=None):
        transfer = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            transfer = receive_transfer(
                transfer, request.user,
                serializer.validated_data.get('items')
            )
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = TransferDetailSerializer(transfer, context={'request': request})
        return Response(output.data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        transfer = self.get_object()
        try:
            transfer = cancel_transfer(transfer, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = TransferDetailSerializer(transfer, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        transfer = self.get_object()
        qs = InventoryTransferHistory.objects.filter(
            transfer=transfer
        ).select_related('performed_by')
        serializer = TransferHistorySerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        tenant_id = _get_tenant_id(request)
        qs = InventoryTransfer.objects.filter(tenant_id=tenant_id).select_related(
            'source_location', 'destination_location', 'created_by'
        )

        fmt = request.query_params.get('export_format', 'csv').lower()
        if fmt == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'Transfer Number', 'Date', 'Source', 'Destination',
                'Type', 'Status', 'Items', 'Created By', 'Created At'
            ])
            for t in qs:
                writer.writerow([
                    t.transfer_number, t.transfer_date,
                    t.source_location.location_name,
                    t.destination_location.location_name,
                    t.transfer_type, t.status,
                    t.items.count(),
                    t.created_by.email if t.created_by else '',
                    t.created_at,
                ])
            output.seek(0)
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="transfers.csv"'
            return response
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Transfers'
            ws.append([
                'Transfer Number', 'Date', 'Source', 'Destination',
                'Type', 'Status', 'Items', 'Created By', 'Created At'
            ])
            for t in qs:
                ws.append([
                    t.transfer_number, str(t.transfer_date),
                    t.source_location.location_name,
                    t.destination_location.location_name,
                    t.transfer_type, t.status,
                    t.items.count(),
                    t.created_by.email if t.created_by else '',
                    str(t.created_at),
                ])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="transfers.xlsx"'
            return response


# ============================================================================
# ADJUSTMENT VIEWSET (Section 7)
# ============================================================================

class AdjustmentReasonViewSet(InventoryFeatureMixin, viewsets.ReadOnlyModelViewSet):
    required_feature = "adjustments"
    """GET /inventory/adjustment-reasons/ — list active adjustment reasons."""
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reason_code', 'reason_name']
    ordering_fields = ['reason_code', 'reason_name', 'adjustment_type']
    ordering = ['reason_name']

    def get_serializer_class(self):
        if self.action == 'list':
            return AdjustmentReasonListSerializer
        return AdjustmentReasonSerializer

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        qs = InventoryAdjustmentReason.objects.for_tenant(tenant_id)
        adjustment_type = self.request.query_params.get('adjustment_type')
        if adjustment_type:
            qs = qs.filter(adjustment_type=adjustment_type)
        show_inactive = self.request.query_params.get('show_inactive', 'false') == 'true'
        if not show_inactive:
            qs = qs.filter(status='ACTIVE')
        return qs.distinct()


class AdjustmentViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    """
    required_feature = "adjustments"
    Stock adjustments with full workflow: create → submit → approve → apply
    Never updates stock directly. Always uses Stock Ledger.
    """
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'status': ['exact'],
        'location': ['exact'],
        'adjustment_type': ['exact'],
        'reason': ['exact'],
        'adjustment_date': ['gte', 'lte'],
    }
    search_fields = ['adjustment_number', 'remarks']
    ordering_fields = ['adjustment_number', 'adjustment_date', 'created_at', 'status']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return AdjustmentListSerializer
        elif self.action in ['create']:
            return CreateAdjustmentSerializer
        return AdjustmentDetailSerializer

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        return InventoryAdjustment.objects.filter(tenant_id=tenant_id).select_related(
            'location', 'reason', 'created_by'
        ).prefetch_related('items__item', 'items__unit')

    def get_permissions(self):
        perms_map = {
            'create': [CanCreateAdjustments],
            'update': [CanEditAdjustments],
            'partial_update': [CanEditAdjustments],
            'destroy': [CanDeleteAdjustments],
            'submit': [CanSubmitAdjustment],
            'approve': [CanApproveAdjustment],
            'reject': [CanApproveAdjustment],
            'apply': [CanApplyAdjustment],
            'export': [CanExportAdjustments],
        }
        for action, perms in perms_map.items():
            if self.action == action:
                return [p() for p in perms]
        return [CanViewAdjustments()]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        tenant_id = _get_tenant_id(request)
        try:
            adjustment = create_adjustment(tenant_id, serializer.validated_data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output_serializer = AdjustmentDetailSerializer(
            adjustment, context={'request': request}
        )
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        adjustment = self.get_object()
        try:
            adjustment = update_adjustment(adjustment, request.data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = AdjustmentDetailSerializer(adjustment, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        adjustment = self.get_object()
        try:
            adjustment = submit_adjustment(adjustment, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = AdjustmentDetailSerializer(adjustment, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        adjustment = self.get_object()
        notes = request.data.get('notes', '')
        try:
            adjustment = approve_adjustment(adjustment, request.user, notes)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = AdjustmentDetailSerializer(adjustment, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        adjustment = self.get_object()
        notes = request.data.get('notes', '')
        try:
            adjustment = reject_adjustment(adjustment, request.user, notes)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = AdjustmentDetailSerializer(adjustment, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def apply(self, request, pk=None):
        adjustment = self.get_object()
        try:
            adjustment = apply_adjustment(adjustment, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = AdjustmentDetailSerializer(adjustment, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        adjustment = self.get_object()
        try:
            adjustment = cancel_adjustment(adjustment, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = AdjustmentDetailSerializer(adjustment, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        adjustment = self.get_object()
        qs = InventoryAdjustmentHistory.objects.filter(
            adjustment=adjustment
        ).select_related('performed_by')
        serializer = AdjustmentHistorySerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        tenant_id = _get_tenant_id(request)
        qs = InventoryAdjustment.objects.filter(tenant_id=tenant_id).select_related(
            'location', 'reason', 'created_by'
        )

        fmt = request.query_params.get('export_format', 'csv').lower()
        if fmt == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'Adjustment Number', 'Date', 'Location', 'Type',
                'Reason', 'Status', 'Items', 'Created By', 'Created At'
            ])
            for a in qs:
                writer.writerow([
                    a.adjustment_number, a.adjustment_date,
                    a.location.location_name,
                    a.get_adjustment_type_display(),
                    a.reason.reason_name, a.status,
                    a.items.count(),
                    a.created_by.email if a.created_by else '',
                    a.created_at,
                ])
            output.seek(0)
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="adjustments.csv"'
            return response
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Adjustments'
            ws.append([
                'Adjustment Number', 'Date', 'Location', 'Type',
                'Reason', 'Status', 'Items', 'Created By', 'Created At'
            ])
            for a in qs:
                ws.append([
                    a.adjustment_number, str(a.adjustment_date),
                    a.location.location_name,
                    a.get_adjustment_type_display(),
                    a.reason.reason_name, a.status,
                    a.items.count(),
                    a.created_by.email if a.created_by else '',
                    str(a.created_at),
                ])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="adjustments.xlsx"'
            return response


# ============================================================================
# LOCATION VIEWSET
# ============================================================================

class LocationViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    required_feature = "locations"
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'status': ['exact'],
        'location_type': ['exact'],
        'city': ['exact', 'icontains'],
        'state': ['exact', 'icontains'],
        'country': ['exact', 'icontains'],
    }
    search_fields = ['location_code', 'location_name', 'city', 'state', 'country']
    ordering_fields = ['location_code', 'location_name', 'city', 'status', 'created_at']
    ordering = ['location_name']

    def get_serializer_class(self):
        if self.action == 'list':
            return LocationListSerializer
        elif self.action == 'tree':
            return LocationTreeSerializer
        return LocationSerializer

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        qs = InventoryLocation.objects.filter(tenant_id=tenant_id).select_related(
            'location_type', 'parent_location'
        )

        show_archived = self.request.query_params.get('show_archived', 'false') == 'true'
        if not show_archived and self.action != 'tree':
            qs = qs.exclude(status='ARCHIVED')

        return qs

    def get_permissions(self):
        perms_map = {
            'create': [CanCreateLocations],
            'update': [CanEditLocations],
            'partial_update': [CanEditLocations],
            'destroy': [CanDeleteLocations],
        }
        for action, perms in perms_map.items():
            if self.action == action:
                return [p() for p in perms]
        return [CanViewLocations()]

    def perform_create(self, serializer):
        user = self.request.user
        serializer.save(
            tenant_id=_get_tenant_id(self.request),
            created_by=user,
            updated_by=user,
        )

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    def perform_destroy(self, instance):
        instance.status = 'ARCHIVED'
        instance.save()

    @action(detail=True, methods=['post'])
    def archive(self, request, pk=None):
        location = self.get_object()
        location.status = 'ARCHIVED'
        location.save()
        return Response({
            'success': True,
            'message': f"Location '{location.location_name}' archived."
        })

    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        location = self.get_object()
        location.status = 'ACTIVE'
        location.save()
        return Response({
            'success': True,
            'message': f"Location '{location.location_name}' restored."
        })

    @action(detail=False, methods=['get'])
    def tree(self, request):
        tenant_id = _get_tenant_id(request)
        locations = InventoryLocation.objects.filter(
            tenant_id=tenant_id,
            parent_location__isnull=True,
        )
        show_archived = request.query_params.get('show_archived', 'false') == 'true'
        if not show_archived:
            locations = locations.exclude(status='ARCHIVED')

        serializer = LocationTreeSerializer(
            locations, many=True,
            context={'request': request, 'filters': {'status': None}}
        )
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        tenant_id = _get_tenant_id(request)
        qs = InventoryLocation.objects.filter(tenant_id=tenant_id)

        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        fmt = request.query_params.get('export_format', 'csv').lower()
        if fmt == 'csv':
            return self._export_csv(qs)
        else:
            return self._export_excel(qs)

    def _export_csv(self, qs):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            'Location Code', 'Location Name', 'Type', 'Parent',
            'City', 'State', 'Country', 'Status'
        ])
        for loc in qs:
            writer.writerow([
                loc.location_code,
                loc.location_name,
                loc.location_type.type_name if loc.location_type else '',
                loc.parent_location.location_code if loc.parent_location else '',
                loc.city, loc.state, loc.country, loc.status,
            ])
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="locations.csv"'
        return response

    def _export_excel(self, qs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Locations'
        ws.append([
            'Location Code', 'Location Name', 'Type', 'Parent',
            'City', 'State', 'Country', 'Status'
        ])
        for loc in qs:
            ws.append([
                loc.location_code,
                loc.location_name,
                loc.location_type.type_name if loc.location_type else '',
                loc.parent_location.location_code if loc.parent_location else '',
                loc.city, loc.state, loc.country, loc.status,
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="locations.xlsx"'
        return response

    @action(detail=False, methods=['post'], url_path='import')
    def import_bulk(self, request):
        """Bulk import locations from CSV/Excel or JSON."""
        file = request.FILES.get('file')
        tenant_id = _get_tenant_id(request)
        created = []
        errors = []

        try:
            if file:
                if file.name.endswith('.csv'):
                    reader = csv.DictReader(io.StringIO(file.read().decode('utf-8-sig')))
                    rows = list(reader)
                elif file.name.endswith(('.xlsx', '.xls')):
                    wb = openpyxl.load_workbook(file, read_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in next(ws.iter_rows())]
                    rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers, row)))
                else:
                    return Response({'error': 'Unsupported file format.'}, status=status.HTTP_400_BAD_REQUEST)
            else:
                serializer = BulkImportSerializer(data=request.data)
                serializer.is_valid(raise_exception=True)
                rows = serializer.validated_data['items']

            for idx, row in enumerate(rows):
                code = str(row.get('Location Code', row.get('location_code', ''))).strip()
                name = str(row.get('Location Name', row.get('location_name', ''))).strip()
                type_name = str(row.get('Type', row.get('location_type', ''))).strip()

                if not code or not name:
                    errors.append({'index': idx, 'error': 'Location Code and Name are required.'})
                    continue

                if InventoryLocation.objects.filter(tenant_id=tenant_id, location_code=code).exists():
                    errors.append({'index': idx, 'error': f"Code '{code}' already exists."})
                    continue

                location_type = None
                if type_name:
                    location_type = InventoryLocationType.objects.filter(
                        tenant_id=tenant_id,
                        type_name__iexact=type_name
                    ).first()

                parent_code = str(row.get('Parent', row.get('parent_location', ''))).strip()
                parent = None
                if parent_code:
                    parent = InventoryLocation.objects.filter(
                        tenant_id=tenant_id, location_code=parent_code
                    ).first()

                loc = InventoryLocation(
                    tenant_id=tenant_id,
                    location_code=code,
                    location_name=name,
                    location_type=location_type,
                    parent_location=parent,
                    city=str(row.get('City', row.get('city', ''))).strip(),
                    state=str(row.get('State', row.get('state', ''))).strip(),
                    country=str(row.get('Country', row.get('country', ''))).strip(),
                    status='ACTIVE',
                )
                loc.save()
                created.append(LocationListSerializer(loc).data)

        except Exception as e:
            return Response({'error': f'Failed to parse file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'created_count': len(created),
            'error_count': len(errors),
            'created': created,
            'errors': errors,
        }, status=status.HTTP_201_CREATED)


# ============================================================================
# RESERVATION VIEWSET (Section 8)
# ============================================================================

class ReservationReasonViewSet(InventoryFeatureMixin, viewsets.ReadOnlyModelViewSet):
    required_feature = "reservations"
    """GET /inventory/reservation-reasons/ — list active reservation reasons."""
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reason_code', 'reason_name']
    ordering_fields = ['reason_code', 'reason_name']
    ordering = ['reason_name']

    def get_serializer_class(self):
        if self.action == 'list':
            return ReservationReasonListSerializer
        return ReservationReasonSerializer

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        qs = InventoryReservationReason.objects.for_tenant(tenant_id)
        show_inactive = self.request.query_params.get('show_inactive', 'false') == 'true'
        if not show_inactive:
            qs = qs.filter(status='ACTIVE')
        return qs.distinct()


class ReservationViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    """
    required_feature = "reservations"
    Stock reservations with full workflow: create → activate → fulfill/cancel/expire
    Never updates stock directly. Always uses Stock Ledger.
    """
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'status': ['exact'],
        'source_location': ['exact'],
        'reservation_type': ['exact'],
        'priority': ['exact'],
        'reservation_date': ['gte', 'lte'],
        'expiry_date': ['gte', 'lte'],
    }
    search_fields = ['reservation_number', 'customer_name', 'reference_number', 'remarks']
    ordering_fields = ['reservation_number', 'reservation_date', 'expiry_date', 'created_at', 'status', 'priority']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return ReservationListSerializer
        elif self.action in ['create']:
            return CreateReservationSerializer
        elif self.action in ['fulfill']:
            return FulfillReservationSerializer
        return ReservationDetailSerializer

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        return InventoryReservation.objects.filter(tenant_id=tenant_id).select_related(
            'source_location', 'reason', 'created_by'
        ).prefetch_related('items__item', 'items__unit')

    def get_permissions(self):
        perms_map = {
            'create': [CanCreateReservations],
            'update': [CanEditReservations],
            'partial_update': [CanEditReservations],
            'destroy': [CanDeleteReservations],
            'activate': [CanActivateReservations],
            'fulfill': [CanFulfillReservations],
            'cancel': [CanCancelReservations],
            'expire': [CanCancelReservations],
            'bulk_cancel': [CanCancelReservations],
            'export': [CanExportReservations],
        }
        for action, perms in perms_map.items():
            if self.action == action:
                return [p() for p in perms]
        return [CanViewReservations()]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        tenant_id = _get_tenant_id(request)
        try:
            reservation = create_reservation(tenant_id, serializer.validated_data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output_serializer = ReservationDetailSerializer(
            reservation, context={'request': request}
        )
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        reservation = self.get_object()
        try:
            reservation = update_reservation(reservation, request.data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = ReservationDetailSerializer(reservation, context={'request': request})
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        reservation = self.get_object()
        if reservation.status != 'DRAFT':
            return Response(
                {'error': f"Cannot delete a reservation in status '{reservation.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        """Activate a reservation — creates RESERVATION ledger entries."""
        reservation = self.get_object()
        try:
            reservation = activate_reservation(reservation, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = ReservationDetailSerializer(reservation, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def fulfill(self, request, pk=None):
        """Fulfill a reservation — releases reserved stock."""
        reservation = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            reservation = fulfill_reservation(
                reservation, request.user,
                serializer.validated_data.get('items')
            )
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = ReservationDetailSerializer(reservation, context={'request': request})
        return Response(output.data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel a reservation — releases reserved stock."""
        reservation = self.get_object()
        try:
            reservation = cancel_reservation(reservation, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = ReservationDetailSerializer(reservation, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def expire(self, request, pk=None):
        """Expire a reservation — releases reserved stock."""
        reservation = self.get_object()
        try:
            reservation = expire_reservation(reservation, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = ReservationDetailSerializer(reservation, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        """GET /inventory/reservations/{id}/history/ — return complete audit trail."""
        reservation = self.get_object()
        qs = InventoryReservationHistory.objects.filter(
            reservation=reservation
        ).select_related('performed_by')
        serializer = ReservationHistorySerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='bulk-cancel')
    def bulk_cancel(self, request):
        """POST /inventory/reservations/bulk-cancel/ — cancel multiple reservations."""
        reservation_ids = request.data.get('reservation_ids', [])
        if not reservation_ids:
            return Response(
                {'error': 'reservation_ids is required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        tenant_id = _get_tenant_id(request)
        results = bulk_cancel_reservations(tenant_id, reservation_ids, request.user)
        return Response({'results': results})

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """Export reservations to CSV/Excel."""
        tenant_id = _get_tenant_id(request)
        qs = InventoryReservation.objects.filter(tenant_id=tenant_id).select_related(
            'source_location', 'reason', 'created_by'
        )

        fmt = request.query_params.get('export_format', 'csv').lower()
        if fmt == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'Reservation Number', 'Date', 'Expiry', 'Location',
                'Type', 'Priority', 'Status', 'Customer',
                'Reference', 'Items', 'Created By', 'Created At'
            ])
            for r in qs:
                writer.writerow([
                    r.reservation_number, r.reservation_date,
                    r.expiry_date or '',
                    r.source_location.location_name,
                    r.get_reservation_type_display(),
                    r.priority, r.status,
                    r.customer_name, r.reference_number,
                    r.items.count(),
                    r.created_by.email if r.created_by else '',
                    r.created_at,
                ])
            output.seek(0)
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="reservations.csv"'
            return response
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Reservations'
            ws.append([
                'Reservation Number', 'Date', 'Expiry', 'Location',
                'Type', 'Priority', 'Status', 'Customer',
                'Reference', 'Items', 'Created By', 'Created At'
            ])
            for r in qs:
                ws.append([
                    r.reservation_number, str(r.reservation_date),
                    str(r.expiry_date) if r.expiry_date else '',
                    r.source_location.location_name,
                    r.get_reservation_type_display(),
                    r.priority, r.status,
                    r.customer_name, r.reference_number,
                    r.items.count(),
                    r.created_by.email if r.created_by else '',
                    str(r.created_at),
                ])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="reservations.xlsx"'
            return response


# ============================================================================
# STOCK COUNT VIEWSET (Section 9)
# ============================================================================

class StockCountReasonViewSet(InventoryFeatureMixin, viewsets.ReadOnlyModelViewSet):
    required_feature = "physical-stock-count"
    """GET /inventory/stock-count-reasons/ — list active stock count reasons."""
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reason_code', 'reason_name']
    ordering_fields = ['reason_code', 'reason_name']
    ordering = ['reason_name']

    def get_serializer_class(self):
        if self.action == 'list':
            return StockCountReasonListSerializer
        return StockCountReasonSerializer

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        qs = StockCountReason.objects.for_tenant(tenant_id)
        show_inactive = self.request.query_params.get('show_inactive', 'false') == 'true'
        if not show_inactive:
            qs = qs.filter(status='ACTIVE')
        return qs.distinct()


class StockCountViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    """
    required_feature = "physical-stock-count"
    Physical Stock Count (Cycle Count) with full workflow:

    DRAFT → ASSIGNED → IN_PROGRESS → SUBMITTED → APPROVED → COMPLETED
                                                           → CANCELLED

    On COMPLETED, differences auto-generate Adjustments via Section 7 services.
    """
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'status': ['exact'],
        'location': ['exact'],
        'count_type': ['exact'],
        'reason': ['exact'],
        'count_date': ['gte', 'lte'],
        'category': ['exact'],
    }
    search_fields = ['count_number', 'remarks']
    ordering_fields = ['count_number', 'count_date', 'created_at', 'status', 'count_type']
    ordering = ['-created_at']



    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        return InventoryStockCount.objects.filter(tenant_id=tenant_id).select_related(
            'location', 'reason', 'category', 'created_by'
        ).prefetch_related('items__item', 'items__unit', 'assigned_counters')

    def get_permissions(self):
        perms_map = {
            'create': [CanCreateStockCounts],
            'update': [CanEditStockCounts],
            'partial_update': [CanEditStockCounts],
            'destroy': [CanEditStockCounts],
            'assign_counters': [CanAssignStockCounts],
            'start': [CanCountStockItems],
            'save_progress': [CanCountStockItems],
            'submit': [CanSubmitStockCounts],
            'approve': [CanApproveStockCounts],
            'complete': [CanApproveStockCounts],
            'cancel': [CanCancelStockCounts],
            'export': [CanExportStockCounts],
            'print_count_sheet': [CanPrintStockCounts],
        }
        for action, perms in perms_map.items():
            if self.action == action:
                return [p() for p in perms]
        return [CanViewStockCounts()]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        tenant_id = _get_tenant_id(request)
        try:
            stock_count = create_stock_count(tenant_id, serializer.validated_data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output_serializer = StockCountDetailSerializer(
            stock_count, context={'request': request}
        )
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)

    def get_serializer_class(self):
        if self.action == 'list':
            return StockCountListSerializer
        elif self.action in ['create']:
            return CreateStockCountSerializer
        elif self.action in ['update', 'partial_update']:
            return UpdateStockCountSerializer
        elif self.action in ['save_progress']:
            return SaveCountProgressSerializer
        elif self.action in ['assign_counters']:
            return AssignCountersSerializer
        elif self.action in ['difference_summary']:
            return StockCountDifferenceSerializer
        return StockCountDetailSerializer

    def update(self, request, *args, **kwargs):
        stock_count = self.get_object()
        if stock_count.status not in ('DRAFT', 'ASSIGNED'):
            return Response(
                {'error': f"Cannot edit a count in status '{stock_count.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = self.get_serializer(stock_count, data=request.data, partial=False)
        serializer.is_valid(raise_exception=True)
        try:
            stock_count = update_stock_count(stock_count, serializer.validated_data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = StockCountDetailSerializer(stock_count, context={'request': request})
        return Response(output.data)

    def partial_update(self, request, *args, **kwargs):
        stock_count = self.get_object()
        if stock_count.status not in ('DRAFT', 'ASSIGNED'):
            return Response(
                {'error': f"Cannot edit a count in status '{stock_count.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = self.get_serializer(stock_count, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        try:
            stock_count = update_stock_count(stock_count, serializer.validated_data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = StockCountDetailSerializer(stock_count, context={'request': request})
        return Response(output.data)

    def destroy(self, request, *args, **kwargs):
        stock_count = self.get_object()
        if stock_count.status != 'DRAFT':
            return Response(
                {'error': f"Cannot delete a count in status '{stock_count.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'], url_path='assign-counters')
    def assign_counters(self, request, pk=None):
        """Assign counters to the stock count."""
        stock_count = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            stock_count = assign_counters(
                stock_count, request.user,
                serializer.validated_data['assigned_counters']
            )
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = StockCountDetailSerializer(stock_count, context={'request': request})
        return Response(output.data)

    @action(detail=True, methods=['post'])
    def start(self, request, pk=None):
        """Start counting — transition ASSIGNED → IN_PROGRESS."""
        stock_count = self.get_object()
        try:
            stock_count = start_counting(stock_count, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = StockCountDetailSerializer(stock_count, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='save-progress')
    def save_progress(self, request, pk=None):
        """Save counting progress — record counted quantities."""
        stock_count = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            stock_count = save_count_progress(
                stock_count, request.user,
                serializer.validated_data['items']
            )
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = StockCountDetailSerializer(stock_count, context={'request': request})
        return Response(output.data)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit stock count for approval."""
        stock_count = self.get_object()
        try:
            stock_count = submit_stock_count(stock_count, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = StockCountDetailSerializer(stock_count, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve a submitted stock count."""
        stock_count = self.get_object()
        notes = request.data.get('notes', '')
        try:
            stock_count = approve_stock_count(stock_count, request.user, notes)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = StockCountDetailSerializer(stock_count, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """
        Complete the stock count — generates Adjustment records from differences.
        The adjustment is auto-approved and applied to update stock ledger.
        """
        stock_count = self.get_object()
        try:
            stock_count = complete_stock_count(stock_count, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = StockCountDetailSerializer(stock_count, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel a stock count (DRAFT, ASSIGNED, or IN_PROGRESS only)."""
        stock_count = self.get_object()
        try:
            stock_count = cancel_stock_count(stock_count, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = StockCountDetailSerializer(stock_count, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='difference-summary')
    def difference_summary(self, request, pk=None):
        """Get the difference summary for a stock count."""
        stock_count = self.get_object()
        summary = get_difference_summary(stock_count)
        return Response(summary)

    @action(detail=True, methods=['get'], url_path='lookup-barcode')
    def lookup_barcode(self, request, pk=None):
        stock_count = self.get_object()
        barcode = request.query_params.get('barcode', '').strip()

        if not barcode:
            return Response(
                {'error': 'barcode query parameter is required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        count_item = stock_count.items.select_related(
                'item', 'item__unit'
            ).filter(scanned_barcode=barcode).first()
        if not count_item:
            return Response(
                {'error': f'Barcode "{barcode}" not found in this stock count.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        diff = count_item.difference_quantity
        if diff == 0 and count_item.counted_quantity is not None:
            status_label = 'MATCH'
        elif diff > 0:
            status_label = 'SURPLUS'
        elif diff < 0:
            status_label = 'SHORTAGE'
        else:
            status_label = 'UNCOUNTED'

        return Response({
            'item_id': count_item.item_id,
            'item_code': count_item.item.item_code,
            'item_name': count_item.item.item_name,
            'count_item_id': count_item.id,
            'expected_quantity': count_item.expected_quantity,
            'counted_quantity': count_item.counted_quantity,
            'difference_quantity': count_item.difference_quantity,
            'scanned_barcode': count_item.scanned_barcode,
            'status': status_label,
            'location_name': stock_count.location.location_name,
            'unit_name': count_item.item.unit.unit_name if count_item.item.unit else '',
        })

    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        """Get the audit trail for a stock count."""
        stock_count = self.get_object()
        qs = InventoryStockCountHistory.objects.filter(
            count=stock_count
        ).select_related('performed_by')
        serializer = StockCountHistorySerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='reload-items')
    def reload_items(self, request, pk=None):
        """Reload items from stock ledger into the count."""
        stock_count = self.get_object()
        try:
            added = reload_items_from_ledger(stock_count, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({
            'success': True,
            'message': f"{added} new item(s) loaded from stock ledger.",
            'items_added': added,
        })

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """Export stock counts to CSV/Excel."""
        tenant_id = _get_tenant_id(request)
        qs = InventoryStockCount.objects.filter(tenant_id=tenant_id).select_related(
            'location', 'reason', 'created_by'
        )

        fmt = request.query_params.get('export_format', 'csv').lower()
        if fmt == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'Count Number', 'Date', 'Type', 'Location',
                'Reason', 'Status',
                'Total Items', 'Items Counted', 'Items with Diff',
                'Created By', 'Created At'
            ])
            for c in qs:
                writer.writerow([
                    c.count_number, c.count_date,
                    c.get_count_type_display(),
                    c.location.location_name,
                    c.reason.reason_name,
                    c.get_status_display(),
                    c.items.count(),
                    c.total_items_counted,
                    c.total_items_with_difference,
                    c.created_by.email if c.created_by else '',
                    c.created_at,
                ])
            output.seek(0)
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="stock_counts.csv"'
            return response
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Stock Counts'
            ws.append([
                'Count Number', 'Date', 'Type', 'Location',
                'Reason', 'Status',
                'Total Items', 'Items Counted', 'Items with Diff',
                'Created By', 'Created At'
            ])
            for c in qs:
                ws.append([
                    c.count_number, str(c.count_date),
                    c.get_count_type_display(),
                    c.location.location_name,
                    c.reason.reason_name,
                    c.get_status_display(),
                    c.items.count(),
                    c.total_items_counted,
                    c.total_items_with_difference,
                    c.created_by.email if c.created_by else '',
                    str(c.created_at),
                ])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="stock_counts.xlsx"'
            return response

    @action(detail=True, methods=['get', 'post', 'delete'], url_path='attachments')
    def attachments(self, request, pk=None):
        """Manage stock count attachments."""
        stock_count = self.get_object()
        from inventory.serializers import StockCountAttachmentSerializer

        if request.method == 'GET':
            qs = InventoryStockCountAttachment.objects.filter(count=stock_count)
            serializer = StockCountAttachmentSerializer(qs, many=True)
            return Response(serializer.data)

        elif request.method == 'POST':
            serializer = StockCountAttachmentSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save(count=stock_count)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'DELETE':
            attachment_id = request.query_params.get('attachment_id')
            if not attachment_id:
                return Response(
                    {'error': 'attachment_id query parameter is required.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            try:
                attachment = InventoryStockCountAttachment.objects.get(
                    id=attachment_id, count=stock_count
                )
                attachment.delete()
                return Response({'success': True, 'message': 'Attachment deleted.'})
            except InventoryStockCountAttachment.DoesNotExist:
                return Response(
                    {'error': 'Attachment not found.'},
                    status=status.HTTP_404_NOT_FOUND,
                )

    @action(detail=True, methods=['get'], url_path='print')
    def print_count_sheet(self, request, pk=None):
        """
        Print a stock count sheet as PDF.
        Requires weasyprint to be installed.
        """
        stock_count = self.get_object()
        items = stock_count.items.select_related('item', 'item__unit').all()

        html_content = render_to_string('stock_count_sheet.html', {
            'stock_count': stock_count,
            'items': items,
        })

        pdf_file = HTML(string=html_content).write_pdf()

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="stock_count_{stock_count.count_number}.pdf"'
        return response



# ============================================================================
# PURCHASE ORDER VIEWSET (Section 10)
# ============================================================================

class PurchaseOrderViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    """
    required_feature = "purchase-orders"
    Purchase Orders with full workflow:

    DRAFT → SENT → PARTIALLY_RECEIVED → RECEIVED → CLOSED
                                              → CANCELLED

    On RECEIVED, stock is updated via Stock Ledger (PURCHASE_IN entries).
    """
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'status': ['exact'],
        'supplier': ['exact'],
        'location': ['exact'],
        'order_date': ['gte', 'lte'],
        'expected_delivery_date': ['gte', 'lte'],
    }
    search_fields = ['order_number', 'supplier_name', 'supplier_reference', 'notes']
    ordering_fields = ['order_number', 'order_date', 'created_at', 'status', 'total_amount']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return PurchaseOrderListSerializer
        elif self.action in ['create']:
            return CreatePurchaseOrderSerializer
        elif self.action in ['update', 'partial_update']:
            return UpdatePurchaseOrderSerializer
        elif self.action in ['receive']:
            return ReceivePurchaseSerializer
        return PurchaseOrderDetailSerializer

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        return PurchaseOrder.objects.filter(tenant_id=tenant_id).select_related(
            'supplier', 'location', 'created_by'
        ).prefetch_related('items__item', 'items__item__unit')

    def get_permissions(self):
        perms_map = {
            'create': [CanCreatePurchaseOrders],
            'update': [CanEditPurchaseOrders],
            'partial_update': [CanEditPurchaseOrders],
            'destroy': [CanDeletePurchaseOrders],
            'send': [CanSendPurchaseOrders],
            'receive': [CanReceivePurchaseOrders],
            'close': [CanClosePurchaseOrders],
            'cancel': [CanCancelPurchaseOrders],
            'export': [CanExportPurchaseOrders],
            'print_po': [CanPrintPurchaseOrders],
            'print_grn': [CanPrintPurchaseOrders],
            'history': [CanViewPurchaseOrders],
            'receipts': [CanViewPurchaseOrders],
        }
        for action, perms in perms_map.items():
            if self.action == action:
                return [p() for p in perms]
        return [CanViewPurchaseOrders()]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        tenant_id = _get_tenant_id(request)
        try:
            purchase_order = create_purchase_order(tenant_id, serializer.validated_data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output_serializer = PurchaseOrderDetailSerializer(
            purchase_order, context={'request': request}
        )
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        purchase_order = self.get_object()
        if purchase_order.status != 'DRAFT':
            return Response(
                {'error': f"Cannot edit a purchase order in status '{purchase_order.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = self.get_serializer(purchase_order, data=request.data, partial=False)
        serializer.is_valid(raise_exception=True)
        try:
            purchase_order = update_purchase_order(purchase_order, serializer.validated_data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = PurchaseOrderDetailSerializer(purchase_order, context={'request': request})
        return Response(output.data)

    def partial_update(self, request, *args, **kwargs):
        purchase_order = self.get_object()
        if purchase_order.status != 'DRAFT':
            return Response(
                {'error': f"Cannot edit a purchase order in status '{purchase_order.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = self.get_serializer(purchase_order, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        try:
            purchase_order = update_purchase_order(purchase_order, serializer.validated_data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = PurchaseOrderDetailSerializer(purchase_order, context={'request': request})
        return Response(output.data)

    def destroy(self, request, *args, **kwargs):
        purchase_order = self.get_object()
        if purchase_order.status != 'DRAFT':
            return Response(
                {'error': f"Cannot delete a purchase order in status '{purchase_order.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def send(self, request, pk=None):
        """Send purchase order to supplier (DRAFT → SENT)."""
        purchase_order = self.get_object()
        try:
            purchase_order = send_purchase_order(purchase_order, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = PurchaseOrderDetailSerializer(purchase_order, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def receive(self, request, pk=None):
        """Receive items against a purchase order."""
        purchase_order = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            purchase_order, receipt = receive_purchase_order_items(
                purchase_order, request.user,
                serializer.validated_data['items']
            )
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = {
            'purchase_order': PurchaseOrderDetailSerializer(
                purchase_order, context={'request': request}
            ).data,
            'receipt': PurchaseReceiptDetailSerializer(
                receipt, context={'request': request}
            ).data,
        }
        return Response(output)

    @action(detail=True, methods=['post'])
    def close(self, request, pk=None):
        """Close a fully received purchase order."""
        purchase_order = self.get_object()
        try:
            purchase_order = close_purchase_order(purchase_order, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = PurchaseOrderDetailSerializer(purchase_order, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel a purchase order (DRAFT, SENT, or PARTIALLY_RECEIVED only)."""
        purchase_order = self.get_object()
        try:
            purchase_order = cancel_purchase_order(purchase_order, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = PurchaseOrderDetailSerializer(purchase_order, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        """GET /inventory/purchase-orders/{id}/history/ — audit trail."""
        purchase_order = self.get_object()
        qs = PurchaseOrderHistory.objects.filter(
            purchase_order=purchase_order
        ).select_related('performed_by')
        serializer = PurchaseOrderHistorySerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def receipts(self, request, pk=None):
        """GET /inventory/purchase-orders/{id}/receipts/ — list receipts."""
        purchase_order = self.get_object()
        qs = PurchaseReceipt.objects.filter(
            purchase_order=purchase_order
        ).select_related('location', 'created_by').prefetch_related('items__item')
        serializer = PurchaseReceiptListSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """Export purchase orders to CSV/Excel."""
        tenant_id = _get_tenant_id(request)
        qs = PurchaseOrder.objects.filter(tenant_id=tenant_id).select_related(
            'supplier', 'location', 'created_by'
        )

        fmt = request.query_params.get('export_format', 'csv').lower()
        if fmt == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'Order Number', 'Order Date', 'Expected Delivery',
                'Supplier', 'Location', 'Status',
                'Subtotal', 'Tax', 'Discount', 'Total',
                'Items', 'Created By', 'Created At'
            ])
            for po in qs:
                writer.writerow([
                    po.order_number, po.order_date,
                    po.expected_delivery_date or '',
                    po.supplier_name or (po.supplier.name if po.supplier else ''),
                    po.location.location_name if po.location else '',
                    po.status,
                    float(po.subtotal), float(po.tax_amount),
                    float(po.discount_amount), float(po.total_amount),
                    po.items.count(),
                    po.created_by.email if po.created_by else '',
                    po.created_at,
                ])
            output.seek(0)
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="purchase_orders.csv"'
            return response
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Purchase Orders'
            ws.append([
                'Order Number', 'Order Date', 'Expected Delivery',
                'Supplier', 'Location', 'Status',
                'Subtotal', 'Tax', 'Discount', 'Total',
                'Items', 'Created By', 'Created At'
            ])
            for po in qs:
                ws.append([
                    po.order_number, str(po.order_date),
                    str(po.expected_delivery_date) if po.expected_delivery_date else '',
                    po.supplier_name or (po.supplier.name if po.supplier else ''),
                    po.location.location_name if po.location else '',
                    po.status,
                    float(po.subtotal), float(po.tax_amount),
                    float(po.discount_amount), float(po.total_amount),
                    po.items.count(),
                    po.created_by.email if po.created_by else '',
                    str(po.created_at),
                ])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="purchase_orders.xlsx"'
            return response

    @action(detail=True, methods=['get'], url_path='print-po')
    def print_po(self, request, pk=None):
        """Print or render a purchase order document."""
        purchase_order = self.get_object()
        from django.http import HttpResponse

        if WEASYPRINT_AVAILABLE:
            try:
                html = render_to_string('inventory/purchase_order.html', {
                    'po': purchase_order,
                    'items': purchase_order.items.select_related('item', 'item__unit').all(),
                    'today': timezone.now().date(),
                })
                pdf = HTML(string=html).write_pdf()
                response = HttpResponse(pdf, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="PO-{purchase_order.order_number}.pdf"'
                return response
            except Exception:
                pass

        # Fallback: render HTML
        html = render_to_string('inventory/purchase_order.html', {
            'po': purchase_order,
            'items': purchase_order.items.select_related('item', 'item__unit').all(),
            'today': timezone.now().date(),
        })
        return HttpResponse(html)

    @action(detail=True, methods=['get'], url_path='print-grn')
    def print_grn(self, request, pk=None):
        """Print a Goods Receipt Note."""
        purchase_order = self.get_object()
        receipt_id = request.query_params.get('receipt_id')
        receipts = PurchaseReceipt.objects.filter(purchase_order=purchase_order)
        if receipt_id:
            receipts = receipts.filter(id=receipt_id)

        receipt = receipts.first()
        if not receipt:
            return Response(
                {'error': 'No receipt found for this purchase order.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        if WEASYPRINT_AVAILABLE:
            try:
                html = render_to_string('inventory/purchase_receipt.html', {
                    'receipt': receipt,
                    'po': purchase_order,
                    'items': receipt.items.select_related('item').all(),
                    'today': timezone.now().date(),
                })
                pdf = HTML(string=html).write_pdf()
                response = HttpResponse(pdf, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="GRN-{receipt.receipt_number}.pdf"'
                return response
            except Exception:
                pass

        html = render_to_string('inventory/purchase_receipt.html', {
            'receipt': receipt,
            'po': purchase_order,
            'items': receipt.items.select_related('item').all(),
            'today': timezone.now().date(),
        })
        return HttpResponse(html)


class PurchaseReceiptViewSet(InventoryFeatureMixin, viewsets.ReadOnlyModelViewSet):
    required_feature = "purchase-orders"
    """Read-only viewset for purchase receipts / GRNs."""
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'purchase_order': ['exact'],
        'location': ['exact'],
        'receipt_date': ['gte', 'lte'],
    }
    search_fields = ['receipt_number', 'notes']
    ordering_fields = ['receipt_number', 'receipt_date', 'created_at']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return PurchaseReceiptListSerializer
        return PurchaseReceiptDetailSerializer

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        return PurchaseReceipt.objects.filter(tenant_id=tenant_id).select_related(
            'purchase_order', 'location', 'created_by'
        ).prefetch_related('items__item')

    def get_permissions(self):
        return [CanViewPurchaseOrders()]


# ============================================================================
# GOODS RECEIPT NOTE VIEWSET (Section 11)
# ============================================================================

class GoodsReceiptViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    """
    required_feature = "goods-receipts"
    Goods Receipt Notes with full workflow:

    DRAFT → PENDING_APPROVAL → APPROVED → RECEIVED → COMPLETED
                                                    → CANCELLED

    On RECEIVED, stock is updated via Stock Ledger (GOODS_RECEIPT entries).
    """
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'status': ['exact'],
        'purchase_order': ['exact'],
        'supplier': ['exact'],
        'location': ['exact'],
        'receipt_date': ['gte', 'lte'],
    }
    search_fields = ['grn_number', 'supplier_name', 'remarks']
    ordering_fields = ['grn_number', 'receipt_date', 'created_at', 'status']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return GRNListSerializer
        elif self.action in ['create']:
            return CreateGRNSerializer
        elif self.action in ['update', 'partial_update']:
            return UpdateGRNSerializer
        return GRNDetailSerializer

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        return InventoryGoodsReceipt.objects.filter(tenant_id=tenant_id).select_related(
            'purchase_order', 'supplier', 'location', 'created_by'
        ).prefetch_related('items__item', 'items__item__unit')

    def get_permissions(self):
        perms_map = {
            'create': [CanCreateGRNs],
            'update': [CanEditGRNs],
            'partial_update': [CanEditGRNs],
            'destroy': [CanDeleteGRNs],
            'submit': [CanSubmitGRN],
            'approve': [CanApproveGRN],
            'receive': [CanReceiveGRN],
            'complete': [CanReceiveGRN],
            'cancel': [CanCancelGRN],
            'export': [CanExportGRNs],
            'print_grn': [CanPrintGRNs],
            'history': [CanViewGRNs],
        }
        for action, perms in perms_map.items():
            if self.action == action:
                return [p() for p in perms]
        return [CanViewGRNs()]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        tenant_id = _get_tenant_id(request)
        try:
            grn = create_grn(tenant_id, serializer.validated_data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output_serializer = GRNDetailSerializer(
            grn, context={'request': request}
        )
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        grn = self.get_object()
        if grn.status != 'DRAFT':
            return Response(
                {'error': f"Cannot edit a GRN in status '{grn.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = self.get_serializer(grn, data=request.data, partial=False)
        serializer.is_valid(raise_exception=True)
        try:
            grn = update_grn(grn, serializer.validated_data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = GRNDetailSerializer(grn, context={'request': request})
        return Response(output.data)

    def partial_update(self, request, *args, **kwargs):
        grn = self.get_object()
        if grn.status != 'DRAFT':
            return Response(
                {'error': f"Cannot edit a GRN in status '{grn.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = self.get_serializer(grn, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        try:
            grn = update_grn(grn, serializer.validated_data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = GRNDetailSerializer(grn, context={'request': request})
        return Response(output.data)

    def destroy(self, request, *args, **kwargs):
        grn = self.get_object()
        if grn.status != 'DRAFT':
            return Response(
                {'error': f"Cannot delete a GRN in status '{grn.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit GRN for approval (DRAFT → PENDING_APPROVAL)."""
        grn = self.get_object()
        try:
            grn = submit_grn(grn, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = GRNDetailSerializer(grn, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve a GRN (PENDING_APPROVAL → APPROVED)."""
        grn = self.get_object()
        notes = request.data.get('notes', '')
        try:
            grn = approve_grn(grn, request.user, notes)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = GRNDetailSerializer(grn, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def receive(self, request, pk=None):
        """Receive goods against an approved GRN (APPROVED → RECEIVED).
        Creates Stock Ledger entries and updates PO quantities."""
        grn = self.get_object()
        try:
            grn = receive_grn(grn, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = GRNDetailSerializer(grn, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Complete a GRN (RECEIVED → COMPLETED)."""
        grn = self.get_object()
        try:
            grn = complete_grn(grn, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = GRNDetailSerializer(grn, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel a GRN (DRAFT, PENDING_APPROVAL, or APPROVED only)."""
        grn = self.get_object()
        try:
            grn = cancel_grn(grn, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = GRNDetailSerializer(grn, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        """GET /inventory/goods-receipts/{id}/history/ — audit trail."""
        grn = self.get_object()
        qs = InventoryGoodsReceiptHistory.objects.filter(
            goods_receipt=grn
        ).select_related('performed_by')
        serializer = GRNHistorySerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """Export GRNs to CSV/Excel."""
        tenant_id = _get_tenant_id(request)
        qs = InventoryGoodsReceipt.objects.filter(tenant_id=tenant_id).select_related(
            'purchase_order', 'supplier', 'location', 'created_by'
        )

        fmt = request.query_params.get('export_format', 'csv').lower()
        if fmt == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'GRN Number', 'Receipt Date', 'PO Number',
                'Supplier', 'Location', 'Status',
                'Items', 'Created By', 'Created At'
            ])
            for g in qs:
                writer.writerow([
                    g.grn_number, g.receipt_date,
                    g.purchase_order.order_number,
                    g.supplier_name or (g.supplier.name if g.supplier else ''),
                    g.location.location_name if g.location else '',
                    g.status,
                    g.items.count(),
                    g.created_by.email if g.created_by else '',
                    g.created_at,
                ])
            output.seek(0)
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="goods_receipts.csv"'
            return response
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Goods Receipts'
            ws.append([
                'GRN Number', 'Receipt Date', 'PO Number',
                'Supplier', 'Location', 'Status',
                'Items', 'Created By', 'Created At'
            ])
            for g in qs:
                ws.append([
                    g.grn_number, str(g.receipt_date),
                    g.purchase_order.order_number,
                    g.supplier_name or (g.supplier.name if g.supplier else ''),
                    g.location.location_name if g.location else '',
                    g.status,
                    g.items.count(),
                    g.created_by.email if g.created_by else '',
                    str(g.created_at),
                ])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="goods_receipts.xlsx"'
            return response

    @action(detail=True, methods=['get'], url_path='print')
    def print_grn(self, request, pk=None):
        """Print a Goods Receipt Note as PDF."""
        grn = self.get_object()

        if WEASYPRINT_AVAILABLE:
            try:
                html = render_to_string('inventory/grn_document.html', {
                    'grn': grn,
                    'items': grn.items.select_related('item', 'item__unit').all(),
                    'today': timezone.now().date(),
                })
                pdf = HTML(string=html).write_pdf()
                response = HttpResponse(pdf, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="GRN-{grn.grn_number}.pdf"'
                return response
            except Exception:
                pass

        html = render_to_string('inventory/grn_document.html', {
            'grn': grn,
            'items': grn.items.select_related('item', 'item__unit').all(),
            'today': timezone.now().date(),
        })
        return HttpResponse(html)


class GoodsReceiptAttachmentViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    required_feature = "goods-receipts"
    """CRUD for GRN attachments."""
    serializer_class = GRNAttachmentSerializer
    permission_classes = [CanEditGRNs]
    filter_backends = [filters.SearchFilter]
    search_fields = ['file_name']

    def get_queryset(self):
        return InventoryGoodsReceiptAttachment.objects.filter(
            goods_receipt_id=self.kwargs.get('goods_receipt_pk')
        )

    def perform_create(self, serializer):
        serializer.save(
            goods_receipt_id=self.kwargs.get('goods_receipt_pk'),
        )


# ============================================================================
# SUPPLIER INVOICE VIEWSET (Section 12)
# ============================================================================

class SupplierInvoiceViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    """
    required_feature = "supplier-invoices"
    Supplier Invoices (Purchase Bills) with full workflow:

    DRAFT → PENDING_APPROVAL → APPROVED → POSTED → PARTIALLY_PAID → PAID
                                                   → CANCELLED
                                                   → VOIDED

    Integrates with GRNs and Payment tracking.
    """
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = {
        'status': ['exact'],
        'payment_status': ['exact'],
        'supplier': ['exact'],
        'purchase_order': ['exact'],
        'invoice_date': ['gte', 'lte'],
        'due_date': ['gte', 'lte'],
        'currency': ['exact'],
    }
    search_fields = [
        'invoice_number', 'supplier_name', 'supplier_invoice_number',
        'remarks', 'terms',
    ]
    ordering_fields = [
        'invoice_number', 'invoice_date', 'due_date',
        'created_at', 'grand_total', 'status', 'payment_status',
    ]
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return SupplierInvoiceListSerializer
        elif self.action in ['create']:
            return CreateSupplierInvoiceSerializer
        elif self.action in ['update', 'partial_update']:
            return UpdateSupplierInvoiceSerializer
        elif self.action in ['payment']:
            return RecordPaymentSerializer
        return SupplierInvoiceDetailSerializer

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        return InventorySupplierInvoice.objects.filter(tenant_id=tenant_id).select_related(
            'supplier', 'purchase_order', 'created_by'
        ).prefetch_related('items__item', 'items__item__unit', 'goods_receipts')

    def get_permissions(self):
        perms_map = {
            'create': [CanCreateSupplierInvoices],
            'update': [CanEditSupplierInvoices],
            'partial_update': [CanEditSupplierInvoices],
            'destroy': [CanDeleteSupplierInvoices],
            'submit': [CanSubmitSupplierInvoice],
            'approve': [CanApproveSupplierInvoice],
            'post': [CanPostSupplierInvoice],
            'payment': [CanRecordPaymentSupplierInvoice],
            'cancel': [CanCancelSupplierInvoice],
            'void': [CanVoidSupplierInvoice],
            'export': [CanExportSupplierInvoices],
            'print_invoice': [CanPrintSupplierInvoices],
            'history': [CanViewSupplierInvoices],
        }
        for action, perms in perms_map.items():
            if self.action == action:
                return [p() for p in perms]
        return [CanViewSupplierInvoices()]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        tenant_id = _get_tenant_id(request)
        try:
            invoice = create_invoice(tenant_id, serializer.validated_data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output_serializer = SupplierInvoiceDetailSerializer(
            invoice, context={'request': request}
        )
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        invoice = self.get_object()
        if invoice.status != 'DRAFT':
            return Response(
                {'error': f"Cannot edit an invoice in status '{invoice.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = self.get_serializer(invoice, data=request.data, partial=False)
        serializer.is_valid(raise_exception=True)
        try:
            invoice = update_invoice(invoice, serializer.validated_data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = SupplierInvoiceDetailSerializer(invoice, context={'request': request})
        return Response(output.data)

    def partial_update(self, request, *args, **kwargs):
        invoice = self.get_object()
        if invoice.status != 'DRAFT':
            return Response(
                {'error': f"Cannot edit an invoice in status '{invoice.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = self.get_serializer(invoice, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        try:
            invoice = update_invoice(invoice, serializer.validated_data, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = SupplierInvoiceDetailSerializer(invoice, context={'request': request})
        return Response(output.data)

    def destroy(self, request, *args, **kwargs):
        invoice = self.get_object()
        if invoice.status != 'DRAFT':
            return Response(
                {'error': f"Cannot delete an invoice in status '{invoice.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit invoice for approval (DRAFT → PENDING_APPROVAL)."""
        invoice = self.get_object()
        try:
            invoice = submit_invoice(invoice, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = SupplierInvoiceDetailSerializer(invoice, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve an invoice (PENDING_APPROVAL → APPROVED)."""
        invoice = self.get_object()
        notes = request.data.get('notes', '')
        try:
            invoice = approve_invoice(invoice, request.user, notes)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = SupplierInvoiceDetailSerializer(invoice, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def post(self, request, pk=None):
        """Post an approved invoice (APPROVED → POSTED)."""
        invoice = self.get_object()
        try:
            invoice = post_invoice(invoice, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = SupplierInvoiceDetailSerializer(invoice, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def payment(self, request, pk=None):
        """Record a payment against an invoice."""
        invoice = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            invoice = record_payment(
                invoice, request.user,
                serializer.validated_data['amount'],
                serializer.validated_data.get('payment_method', 'Bank Transfer'),
                serializer.validated_data.get('payment_date'),
                serializer.validated_data.get('reference', ''),
                serializer.validated_data.get('remarks', ''),
            )
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = SupplierInvoiceDetailSerializer(invoice, context={'request': request})
        return Response(output.data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel an invoice (DRAFT, PENDING_APPROVAL, or APPROVED only)."""
        invoice = self.get_object()
        try:
            invoice = cancel_invoice(invoice, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = SupplierInvoiceDetailSerializer(invoice, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def void(self, request, pk=None):
        """Void a posted/paid invoice (POSTED, PARTIALLY_PAID, or PAID only)."""
        invoice = self.get_object()
        try:
            invoice = void_invoice(invoice, request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        serializer = SupplierInvoiceDetailSerializer(invoice, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        """GET /inventory/supplier-invoices/{id}/history/ — audit trail."""
        invoice = self.get_object()
        qs = InventorySupplierInvoiceHistory.objects.filter(
            invoice=invoice
        ).select_related('performed_by')
        serializer = SupplierInvoiceHistorySerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """Export supplier invoices to CSV/Excel."""
        tenant_id = _get_tenant_id(request)
        qs = InventorySupplierInvoice.objects.filter(tenant_id=tenant_id).select_related(
            'supplier', 'purchase_order', 'created_by'
        )

        fmt = request.query_params.get('export_format', 'csv').lower()
        if fmt == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'Invoice Number', 'Invoice Date', 'Due Date',
                'Supplier', 'Supplier Invoice #', 'Currency',
                'PO Number', 'Status', 'Payment Status',
                'Subtotal', 'Discount', 'Tax',
                'Shipping', 'Other Charges',
                'Grand Total', 'Outstanding',
                'Items', 'Created By', 'Created At'
            ])
            for inv in qs:
                writer.writerow([
                    inv.invoice_number, inv.invoice_date,
                    inv.due_date or '',
                    inv.supplier_name or (inv.supplier.name if inv.supplier else ''),
                    inv.supplier_invoice_number,
                    inv.currency,
                    inv.purchase_order.order_number if inv.purchase_order else '',
                    inv.status, inv.payment_status,
                    float(inv.subtotal), float(inv.discount_amount),
                    float(inv.tax_amount),
                    float(inv.shipping_charges), float(inv.other_charges),
                    float(inv.grand_total), float(inv.outstanding_amount),
                    inv.items.count(),
                    inv.created_by.email if inv.created_by else '',
                    inv.created_at,
                ])
            output.seek(0)
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="supplier_invoices.csv"'
            return response
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Supplier Invoices'
            ws.append([
                'Invoice Number', 'Invoice Date', 'Due Date',
                'Supplier', 'Supplier Invoice #', 'Currency',
                'PO Number', 'Status', 'Payment Status',
                'Subtotal', 'Discount', 'Tax',
                'Shipping', 'Other Charges',
                'Grand Total', 'Outstanding',
                'Items', 'Created By', 'Created At'
            ])
            for inv in qs:
                ws.append([
                    inv.invoice_number, str(inv.invoice_date),
                    str(inv.due_date) if inv.due_date else '',
                    inv.supplier_name or (inv.supplier.name if inv.supplier else ''),
                    inv.supplier_invoice_number,
                    inv.currency,
                    inv.purchase_order.order_number if inv.purchase_order else '',
                    inv.status, inv.payment_status,
                    float(inv.subtotal), float(inv.discount_amount),
                    float(inv.tax_amount),
                    float(inv.shipping_charges), float(inv.other_charges),
                    float(inv.grand_total), float(inv.outstanding_amount),
                    inv.items.count(),
                    inv.created_by.email if inv.created_by else '',
                    str(inv.created_at),
                ])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="supplier_invoices.xlsx"'
            return response

    @action(detail=True, methods=['get'], url_path='print')
    def print_invoice(self, request, pk=None):
        """Print a supplier invoice as PDF."""
        invoice = self.get_object()

        if WEASYPRINT_AVAILABLE:
            try:
                html = render_to_string('inventory/supplier_invoice.html', {
                    'invoice': invoice,
                    'items': invoice.items.select_related('item', 'item__unit').all(),
                    'today': timezone.now().date(),
                })
                pdf = HTML(string=html).write_pdf()
                response = HttpResponse(pdf, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="SI-{invoice.invoice_number}.pdf"'
                return response
            except Exception:
                pass

        html = render_to_string('inventory/supplier_invoice.html', {
            'invoice': invoice,
            'items': invoice.items.select_related('item', 'item__unit').all(),
            'today': timezone.now().date(),
        })
        return HttpResponse(html)


class SupplierInvoiceAttachmentViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    required_feature = "supplier-invoices"
    """CRUD for supplier invoice attachments."""
    serializer_class = SupplierInvoiceAttachmentSerializer
    permission_classes = [CanEditSupplierInvoices]
    filter_backends = [filters.SearchFilter]
    search_fields = ['file_name']

    def get_queryset(self):
        return InventorySupplierInvoiceAttachment.objects.filter(
            invoice_id=self.kwargs.get('supplier_invoice_pk')
        )

    def perform_create(self, serializer):
        serializer.save(
            invoice_id=self.kwargs.get('supplier_invoice_pk'),
        )


# ============================================================================
# PURCHASE RETURN VIEWS (Section 13)
# ============================================================================

class PurchaseReturnViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    """
    required_feature = "purchase-returns"
    ViewSet for Purchase Returns with full workflow support.

    Actions:
      - Standard CRUD (create/update/destroy)
      - submit / approve / reject / return_to_supplier / complete / cancel
      - history / export / print
    """
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'supplier', 'return_date', 'purchase_order', 'goods_receipt', 'supplier_invoice']
    search_fields = ['return_number', 'supplier_name', 'return_reason']
    ordering_fields = ['return_number', 'return_date', 'created_at', 'total_amount']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return PurchaseReturnListSerializer
        if self.action in ['create', 'bulk_create']:
            return CreatePurchaseReturnSerializer
        if self.action in ['update', 'partial_update']:
            return UpdatePurchaseReturnSerializer
        if self.action == 'history':
            return PurchaseReturnHistorySerializer
        return PurchaseReturnDetailSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'history']:
            permission_classes = [CanViewPurchaseReturns]
        elif self.action == 'create':
            permission_classes = [CanCreatePurchaseReturns]
        elif self.action in ['update', 'partial_update']:
            permission_classes = [CanEditPurchaseReturns]
        elif self.action == 'destroy':
            permission_classes = [CanDeletePurchaseReturns]
        elif self.action == 'submit':
            permission_classes = [CanSubmitPurchaseReturn]
        elif self.action in ['approve', 'reject']:
            permission_classes = [CanApprovePurchaseReturn]
        elif self.action == 'return_to_supplier':
            permission_classes = [CanReturnPurchaseReturn]
        elif self.action == 'complete':
            permission_classes = [CanCompletePurchaseReturn]
        elif self.action == 'cancel':
            permission_classes = [CanCancelPurchaseReturn]
        elif self.action == 'export':
            permission_classes = [CanExportPurchaseReturns]
        else:
            permission_classes = [CanViewPurchaseReturns]
        return [p() for p in permission_classes]

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        qs = InventoryPurchaseReturn.objects.filter(tenant_id=tenant_id)
        if self.action == 'list':
            qs = qs.prefetch_related('items')
        else:
            qs = qs.prefetch_related(
                'items', 'items__item', 'items__item__unit',
                'history', 'history__performed_by',
                'attachments',
            )
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        from inventory.services.purchase_return_service import create_return
        tenant_id = _get_tenant_id(request)
        data = serializer.validated_data
        items_data = data.pop('items', [])
        try:
            return_obj = create_return(
                {**data, 'tenant_id': tenant_id},
                items_data,
                user=request.user,
            )
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output_serializer = PurchaseReturnDetailSerializer(
            return_obj, context={'request': request}
        )
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        return_obj = self.get_object()
        if return_obj.status != 'DRAFT':
            return Response(
                {'error': f"Cannot edit a return in status '{return_obj.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = self.get_serializer(return_obj, data=request.data, partial=False)
        serializer.is_valid(raise_exception=True)
        from inventory.services.purchase_return_service import update_return
        data = serializer.validated_data
        items_data = data.pop('items', None)
        try:
            update_return(return_obj, data, items_data, user=request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = PurchaseReturnDetailSerializer(return_obj, context={'request': request})
        return Response(output.data)

    def partial_update(self, request, *args, **kwargs):
        return_obj = self.get_object()
        if return_obj.status != 'DRAFT':
            return Response(
                {'error': f"Cannot edit a return in status '{return_obj.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = self.get_serializer(return_obj, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        from inventory.services.purchase_return_service import update_return
        data = serializer.validated_data
        items_data = data.pop('items', None)
        try:
            update_return(return_obj, data, items_data, user=request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = PurchaseReturnDetailSerializer(return_obj, context={'request': request})
        return Response(output.data)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        from inventory.services.purchase_return_service import submit_return
        return_obj = self.get_object()
        try:
            submit_return(return_obj, user=request.user)
            serializer = self.get_serializer(return_obj)
            return Response(serializer.data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        from inventory.services.purchase_return_service import approve_return
        return_obj = self.get_object()
        notes = request.data.get('notes', '')
        try:
            approve_return(return_obj, user=request.user, notes=notes)
            serializer = self.get_serializer(return_obj)
            return Response(serializer.data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        from inventory.services.purchase_return_service import reject_return
        return_obj = self.get_object()
        notes = request.data.get('notes', '')
        try:
            reject_return(return_obj, user=request.user, notes=notes)
            serializer = self.get_serializer(return_obj)
            return Response(serializer.data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], url_path='return-to-supplier')
    def return_to_supplier(self, request, pk=None):
        from inventory.services.purchase_return_service import return_to_supplier
        return_obj = self.get_object()
        try:
            return_to_supplier(return_obj, user=request.user)
            serializer = self.get_serializer(return_obj)
            return Response(serializer.data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        from inventory.services.purchase_return_service import complete_return
        return_obj = self.get_object()
        try:
            complete_return(return_obj, user=request.user)
            serializer = self.get_serializer(return_obj)
            return Response(serializer.data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        from inventory.services.purchase_return_service import cancel_return
        return_obj = self.get_object()
        remarks = request.data.get('remarks', '')
        try:
            cancel_return(return_obj, user=request.user, remarks=remarks)
            serializer = self.get_serializer(return_obj)
            return Response(serializer.data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        from inventory.services.purchase_return_service import get_history
        return_obj = self.get_object()
        qs = get_history(return_obj)
        serializer = PurchaseReturnHistorySerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def export(self, request):
        tenant_id = _get_tenant_id(request)
        qs = InventoryPurchaseReturn.objects.filter(tenant_id=tenant_id)
        fmt = request.query_params.get('export_format', 'csv').lower()

        if fmt == 'csv':
            import csv, io
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'Return Number', 'Return Date', 'Supplier',
                'Reason', 'Status',
                'Subtotal', 'Tax', 'Total',
                'Items', 'Created By', 'Created At'
            ])
            for r in qs:
                writer.writerow([
                    r.return_number, r.return_date,
                    r.supplier_name or (r.supplier.name if r.supplier else ''),
                    r.return_reason, r.status,
                    float(r.subtotal), float(r.tax_amount), float(r.total_amount),
                    r.items.count(),
                    r.created_by.email if r.created_by else '',
                    r.created_at,
                ])
            output.seek(0)
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="purchase_returns.csv"'
            return response
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Purchase Returns'
            ws.append([
                'Return Number', 'Return Date', 'Supplier',
                'Reason', 'Status',
                'Subtotal', 'Tax', 'Total',
                'Items', 'Created By', 'Created At'
            ])
            for r in qs:
                ws.append([
                    r.return_number, str(r.return_date),
                    r.supplier_name or (r.supplier.name if r.supplier else ''),
                    r.return_reason, r.status,
                    float(r.subtotal), float(r.tax_amount), float(r.total_amount),
                    r.items.count(),
                    r.created_by.email if r.created_by else '',
                    str(r.created_at),
                ])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="purchase_returns.xlsx"'
            return response

    @action(detail=True, methods=['get'], url_path='print')
    def print_return(self, request, pk=None):
        """Print a purchase return as PDF/HTML."""
        return_obj = self.get_object()
        return HttpResponse(f"Purchase Return: {return_obj.return_number}")


class PurchaseReturnAttachmentViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    required_feature = "purchase-returns"
    """CRUD for purchase return attachments."""
    serializer_class = PurchaseReturnAttachmentSerializer
    permission_classes = [CanEditPurchaseReturns]
    filter_backends = [filters.SearchFilter]
    search_fields = ['file_name']

    def get_queryset(self):
        return InventoryPurchaseReturnAttachment.objects.filter(
            purchase_return_id=self.kwargs.get('purchase_return_pk')
        )

    def perform_create(self, serializer):
        serializer.save(
            purchase_return_id=self.kwargs.get('purchase_return_pk'),
        )


# ============================================================================
# SUPPLIER PAYMENT VIEWS (Section 14)
# ============================================================================

class SupplierPaymentViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    """
    required_feature = "supplier-payments"
    ViewSet for Supplier Payments with full workflow support.

    Actions:
      - Standard CRUD (create/update/destroy)
      - submit / approve / post / allocate / complete / cancel / void
      - history / export / reports
    """
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'supplier', 'payment_method', 'payment_date']
    search_fields = ['payment_number', 'supplier_name', 'reference_number']
    ordering_fields = ['payment_number', 'payment_date', 'created_at', 'total_amount']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return SupplierPaymentListSerializer
        if self.action == 'create':
            return CreateSupplierPaymentSerializer
        if self.action in ['update', 'partial_update']:
            return UpdateSupplierPaymentSerializer
        if self.action == 'allocate':
            return AllocatePaymentSerializer
        if self.action == 'history':
            return SupplierPaymentHistorySerializer
        return SupplierPaymentDetailSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'history']:
            permission_classes = [CanViewSupplierPayments]
        elif self.action == 'create':
            permission_classes = [CanCreateSupplierPayments]
        elif self.action in ['update', 'partial_update']:
            permission_classes = [CanEditSupplierPayments]
        elif self.action == 'destroy':
            permission_classes = [CanDeleteSupplierPayments]
        elif self.action == 'submit':
            permission_classes = [CanSubmitSupplierPayment]
        elif self.action == 'approve':
            permission_classes = [CanApproveSupplierPayment]
        elif self.action == 'post':
            permission_classes = [CanPostSupplierPayment]
        elif self.action == 'allocate':
            permission_classes = [CanAllocateSupplierPayment]
        elif self.action == 'cancel':
            permission_classes = [CanCancelSupplierPayment]
        elif self.action == 'void':
            permission_classes = [CanVoidSupplierPayment]
        elif self.action == 'export':
            permission_classes = [CanExportSupplierPayments]
        else:
            permission_classes = [CanViewSupplierPayments]
        return [p() for p in permission_classes]

    def get_queryset(self):
        tenant_id = _get_tenant_id(self.request)
        qs = InventorySupplierPayment.objects.filter(tenant_id=tenant_id)
        if self.action == 'list':
            qs = qs.prefetch_related('allocations')
        else:
            qs = qs.prefetch_related(
                'allocations', 'allocations__supplier_invoice',
                'history', 'history__performed_by',
                'attachments',
            )
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        tenant_id = _get_tenant_id(request)
        from inventory.services.supplier_payment_service import create_payment
        try:
            payment = create_payment(
                {**serializer.validated_data, 'tenant_id': tenant_id},
                user=request.user,
            )
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = SupplierPaymentDetailSerializer(payment, context={'request': request})
        return Response(output.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        payment = self.get_object()
        if payment.status != 'DRAFT':
            return Response(
                {'error': f"Cannot edit a payment in status '{payment.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = self.get_serializer(payment, data=request.data, partial=False)
        serializer.is_valid(raise_exception=True)
        from inventory.services.supplier_payment_service import update_payment
        try:
            update_payment(payment, serializer.validated_data, user=request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = SupplierPaymentDetailSerializer(payment, context={'request': request})
        return Response(output.data)

    def partial_update(self, request, *args, **kwargs):
        payment = self.get_object()
        if payment.status != 'DRAFT':
            return Response(
                {'error': f"Cannot edit a payment in status '{payment.get_status_display()}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = self.get_serializer(payment, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        from inventory.services.supplier_payment_service import update_payment
        try:
            update_payment(payment, serializer.validated_data, user=request.user)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        output = SupplierPaymentDetailSerializer(payment, context={'request': request})
        return Response(output.data)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        from inventory.services.supplier_payment_service import submit_payment
        payment = self.get_object()
        try:
            submit_payment(payment, user=request.user)
            serializer = self.get_serializer(payment)
            return Response(serializer.data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        from inventory.services.supplier_payment_service import approve_payment
        payment = self.get_object()
        notes = request.data.get('notes', '')
        try:
            approve_payment(payment, user=request.user, notes=notes)
            serializer = self.get_serializer(payment)
            return Response(serializer.data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def post(self, request, pk=None):
        from inventory.services.supplier_payment_service import post_payment
        payment = self.get_object()
        try:
            post_payment(payment, user=request.user)
            serializer = self.get_serializer(payment)
            return Response(serializer.data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def allocate(self, request, pk=None):
        from inventory.services.supplier_payment_service import allocate_payment
        payment = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            allocate_payment(payment, serializer.validated_data['allocations'], user=request.user)
            output = SupplierPaymentDetailSerializer(payment, context={'request': request})
            return Response(output.data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        from inventory.services.supplier_payment_service import cancel_payment
        payment = self.get_object()
        remarks = request.data.get('remarks', '')
        try:
            cancel_payment(payment, user=request.user, remarks=remarks)
            serializer = self.get_serializer(payment)
            return Response(serializer.data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def void(self, request, pk=None):
        from inventory.services.supplier_payment_service import void_payment
        payment = self.get_object()
        remarks = request.data.get('remarks', '')
        try:
            void_payment(payment, user=request.user, remarks=remarks)
            serializer = self.get_serializer(payment)
            return Response(serializer.data)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        from inventory.services.supplier_payment_service import get_history
        payment = self.get_object()
        qs = get_history(payment)
        serializer = SupplierPaymentHistorySerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def export(self, request):
        tenant_id = _get_tenant_id(request)
        qs = InventorySupplierPayment.objects.filter(tenant_id=tenant_id)
        fmt = request.query_params.get('export_format', 'csv').lower()

        if fmt == 'csv':
            import csv, io as csv_io
            output = csv_io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'Payment Number', 'Payment Date', 'Supplier',
                'Method', 'Reference', 'Total Amount',
                'Allocated', 'Unallocated', 'Status',
                'Created By', 'Created At'
            ])
            for p in qs:
                writer.writerow([
                    p.payment_number, p.payment_date,
                    p.supplier_name or (p.supplier.name if p.supplier else ''),
                    p.payment_method, p.reference_number,
                    float(p.total_amount), float(p.allocated_amount),
                    float(p.unallocated_amount),
                    p.status,
                    p.created_by.email if p.created_by else '',
                    p.created_at,
                ])
            output.seek(0)
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="supplier_payments.csv"'
            return response
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Supplier Payments'
            ws.append([
                'Payment Number', 'Payment Date', 'Supplier',
                'Method', 'Reference', 'Total Amount',
                'Allocated', 'Unallocated', 'Status',
                'Created By', 'Created At'
            ])
            for p in qs:
                ws.append([
                    p.payment_number, str(p.payment_date),
                    p.supplier_name or (p.supplier.name if p.supplier else ''),
                    p.payment_method, p.reference_number,
                    float(p.total_amount), float(p.allocated_amount),
                    float(p.unallocated_amount),
                    p.status,
                    p.created_by.email if p.created_by else '',
                    str(p.created_at),
                ])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="supplier_payments.xlsx"'
            return response

    @action(detail=False, methods=['get'], url_path='reports/outstanding')
    def outstanding_report(self, request):
        """Outstanding payables report."""
        from inventory.services.supplier_payment_service import get_outstanding_payables
        tenant_id = _get_tenant_id(request)
        supplier_id = request.query_params.get('supplier')
        qs = get_outstanding_payables(tenant_id, supplier_id=supplier_id)
        from inventory.serializers import SupplierInvoiceListSerializer
        serializer = SupplierInvoiceListSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='reports/cash-flow')
    def cash_flow_report(self, request):
        """Cash flow report."""
        from inventory.services.supplier_payment_service import get_cash_flow_report
        tenant_id = _get_tenant_id(request)
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if not date_from or not date_to:
            return Response({'error': 'date_from and date_to are required.'},
                            status=status.HTTP_400_BAD_REQUEST)
        report = get_cash_flow_report(tenant_id, date_from, date_to)
        return Response(report)


class SupplierPaymentAttachmentViewSet(InventoryFeatureMixin, viewsets.ModelViewSet):
    required_feature = "supplier-payments"
    """CRUD for supplier payment attachments."""
    serializer_class = SupplierPaymentAttachmentSerializer
    permission_classes = [CanEditSupplierPayments]
    filter_backends = [filters.SearchFilter]
    search_fields = ['file_name']

    def get_queryset(self):
        return InventorySupplierPaymentAttachment.objects.filter(
            payment_id=self.kwargs.get('supplier_payment_pk')
        )

    def perform_create(self, serializer):
        serializer.save(
            payment_id=self.kwargs.get('supplier_payment_pk'),
        )


# ============================================================================
# DASHBOARD & REPORTS VIEWS (Section 21)
# ============================================================================

class DashboardViewSet(InventoryFeatureMixin, viewsets.GenericViewSet):
    """
    required_feature = "dashboard"
    Read-only ViewSet for the Inventory Dashboard and all reports.

    Endpoints:
      GET /inventory/dashboard/              - Full dashboard data
      GET /inventory/dashboard/cards/        - Summary cards only
      GET /inventory/dashboard/charts/       - Chart data only
      GET /inventory/reports/<report_type>/  - Individual reports
    """
    permission_classes = [CanViewDashboard]
    filter_backends = [filters.SearchFilter]
    search_fields = ['item__item_code', 'item__item_name']

    def _get_tenant(self, request):
        return _get_tenant_id(request)

    def _get_filters(self, request):
        """Extract common filters from query params."""
        filters = {}
        for param in ['item_id', 'location_id', 'category_id', 'brand_id',
                      'supplier_id', 'transaction_type', 'status',
                      'adjustment_type', 'transfer_type', 'priority',
                      'search', 'date_from', 'date_to']:
            val = request.query_params.get(param)
            if val:
                filters[param] = val
        return filters

    # ---- Dashboard ----

    @action(detail=False, methods=['get'])
    def cards(self, request):
        from inventory.services.dashboard_service import get_stock_summary_cards
        data = get_stock_summary_cards(self._get_tenant(request))
        return Response(data)

    @action(detail=False, methods=['get'])
    def charts(self, request):
        from inventory.services.dashboard_service import (
            get_stock_movement_trend, get_monthly_purchase_trend,
            get_catergory_wise_inventory, get_warehouse_wise_inventory,
            get_inventory_aging_distribution,
        )
        tenant_id = self._get_tenant(request)
        days = int(request.query_params.get('days', 30))
        return Response({
            'stock_movement': get_stock_movement_trend(tenant_id, days),
            'monthly_purchases': get_monthly_purchase_trend(tenant_id),
            'category_inventory': get_catergory_wise_inventory(tenant_id),
            'warehouse_inventory': get_warehouse_wise_inventory(tenant_id),
            'aging_distribution': get_inventory_aging_distribution(tenant_id),
        })

    def list(self, request):
        """Full dashboard data."""
        from inventory.services.dashboard_service import (
            get_stock_summary_cards, get_recent_transactions,
            get_recent_transfers, get_recent_adjustments,
            get_recent_purchase_orders, get_recent_goods_receipts,
            get_recent_purchase_returns, get_recent_supplier_invoices,
            low_stock_report, out_of_stock_report,
        )
        from inventory.serializers import (
            StockLedgerSerializer, TransferListSerializer,
            AdjustmentListSerializer, PurchaseOrderListSerializer,
            GRNListSerializer, PurchaseReturnListSerializer,
            SupplierInvoiceListSerializer,
        )
        tenant_id = self._get_tenant(request)
        limit = int(request.query_params.get('limit', 10))
        return Response({
            'cards': get_stock_summary_cards(tenant_id),
            'recent_transactions': StockLedgerSerializer(
                get_recent_transactions(tenant_id, limit), many=True
            ).data,
            'recent_transfers': TransferListSerializer(
                get_recent_transfers(tenant_id, 5), many=True
            ).data,
            'recent_adjustments': AdjustmentListSerializer(
                get_recent_adjustments(tenant_id, 5), many=True
            ).data,
            'recent_purchase_orders': PurchaseOrderListSerializer(
                get_recent_purchase_orders(tenant_id, 5), many=True
            ).data,
            'recent_goods_receipts': GRNListSerializer(
                get_recent_goods_receipts(tenant_id, 5), many=True
            ).data,
            'recent_purchase_returns': PurchaseReturnListSerializer(
                get_recent_purchase_returns(tenant_id, 5), many=True
            ).data,
            'recent_supplier_invoices': SupplierInvoiceListSerializer(
                get_recent_supplier_invoices(tenant_id, 5), many=True
            ).data,
            'low_stock': low_stock_report(tenant_id),
            'out_of_stock': out_of_stock_report(tenant_id),
        })

    # ---- Stock Reports ----

    @action(detail=False, methods=['get'], url_path='reports/current-stock')
    def current_stock(self, request):
        from inventory.services.dashboard_service import current_stock_report
        data = current_stock_report(self._get_tenant(request), self._get_filters(request))
        return Response(data)

    @action(detail=False, methods=['get'], url_path='reports/stock-ledger')
    def stock_ledger(self, request):
        from inventory.services.dashboard_service import stock_ledger_report
        qs = stock_ledger_report(self._get_tenant(request), self._get_filters(request))
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = StockLedgerSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = StockLedgerSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='reports/stock-summary')
    def stock_summary(self, request):
        from inventory.services.dashboard_service import stock_summary_report
        qs = stock_summary_report(self._get_tenant(request), self._get_filters(request))
        from inventory.serializers import StockSummarySerializer
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = StockSummarySerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = StockSummarySerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='reports/stock-movement')
    def stock_movement(self, request):
        from inventory.services.dashboard_service import stock_movement_report
        data = stock_movement_report(self._get_tenant(request), self._get_filters(request))
        return Response(list(data))

    @action(detail=False, methods=['get'], url_path='reports/valuation')
    def valuation(self, request):
        from inventory.services.dashboard_service import inventory_valuation_report
        data = inventory_valuation_report(self._get_tenant(request), self._get_filters(request))
        return Response(data)

    @action(detail=False, methods=['get'], url_path='reports/reserved-stock')
    def reserved_stock(self, request):
        from inventory.services.dashboard_service import reserved_stock_report
        data = reserved_stock_report(self._get_tenant(request), self._get_filters(request))
        return Response(list(data))

    @action(detail=False, methods=['get'], url_path='reports/damaged-stock')
    def damaged_stock(self, request):
        from inventory.services.dashboard_service import damaged_stock_report
        data = damaged_stock_report(self._get_tenant(request), self._get_filters(request))
        return Response(list(data))

    # ---- Operational Reports ----

    @action(detail=False, methods=['get'], url_path='reports/adjustments')
    def adjustments(self, request):
        from inventory.services.dashboard_service import adjustment_report
        qs = adjustment_report(self._get_tenant(request), self._get_filters(request))
        from inventory.serializers import AdjustmentListSerializer
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = AdjustmentListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = AdjustmentListSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='reports/transfers')
    def transfers(self, request):
        from inventory.services.dashboard_service import transfer_report
        qs = transfer_report(self._get_tenant(request), self._get_filters(request))
        from inventory.serializers import TransferListSerializer
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = TransferListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = TransferListSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='reports/reservations')
    def reservations(self, request):
        from inventory.services.dashboard_service import reservation_report
        qs = reservation_report(self._get_tenant(request), self._get_filters(request))
        from inventory.serializers import ReservationListSerializer
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = ReservationListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = ReservationListSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='reports/stock-counts')
    def stock_counts(self, request):
        from inventory.services.dashboard_service import stock_count_report
        qs = stock_count_report(self._get_tenant(request), self._get_filters(request))
        from inventory.serializers import StockCountListSerializer
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = StockCountListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = StockCountListSerializer(qs, many=True)
        return Response(serializer.data)

    # ---- Purchase Reports ----

    @action(detail=False, methods=['get'], url_path='reports/purchase-orders')
    def purchase_orders(self, request):
        from inventory.services.dashboard_service import purchase_order_report
        qs = purchase_order_report(self._get_tenant(request), self._get_filters(request))
        from inventory.serializers import PurchaseOrderListSerializer
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = PurchaseOrderListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = PurchaseOrderListSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='reports/goods-receipts')
    def goods_receipts(self, request):
        from inventory.services.dashboard_service import goods_receipt_report
        qs = goods_receipt_report(self._get_tenant(request), self._get_filters(request))
        from inventory.serializers import GRNListSerializer
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = GRNListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = GRNListSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='reports/supplier-invoices')
    def supplier_invoices(self, request):
        from inventory.services.dashboard_service import supplier_invoice_report
        qs = supplier_invoice_report(self._get_tenant(request), self._get_filters(request))
        from inventory.serializers import SupplierInvoiceListSerializer
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = SupplierInvoiceListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = SupplierInvoiceListSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='reports/purchase-returns')
    def purchase_returns(self, request):
        from inventory.services.dashboard_service import purchase_return_report
        qs = purchase_return_report(self._get_tenant(request), self._get_filters(request))
        from inventory.serializers import PurchaseReturnListSerializer
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = PurchaseReturnListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = PurchaseReturnListSerializer(qs, many=True)
        return Response(serializer.data)

    # ---- Analytics ----

    @action(detail=False, methods=['get'], url_path='reports/low-stock')
    def low_stock(self, request):
        from inventory.services.dashboard_service import low_stock_report
        data = low_stock_report(self._get_tenant(request))
        return Response(data)

    @action(detail=False, methods=['get'], url_path='reports/out-of-stock')
    def out_of_stock(self, request):
        from inventory.services.dashboard_service import out_of_stock_report
        data = out_of_stock_report(self._get_tenant(request))
        return Response(data)

    @action(detail=False, methods=['get'], url_path='reports/reorder')
    def reorder(self, request):
        from inventory.services.dashboard_service import reorder_report
        data = reorder_report(self._get_tenant(request))
        return Response(data)

    @action(detail=False, methods=['get'], url_path='reports/fast-moving')
    def fast_moving(self, request):
        from inventory.services.dashboard_service import fast_moving_items
        data = fast_moving_items(self._get_tenant(request))
        return Response(data)

    @action(detail=False, methods=['get'], url_path='reports/slow-moving')
    def slow_moving(self, request):
        from inventory.services.dashboard_service import slow_moving_items
        data = slow_moving_items(self._get_tenant(request))
        return Response(data)

    @action(detail=False, methods=['get'], url_path='reports/dead-stock')
    def dead_stock(self, request):
        from inventory.services.dashboard_service import dead_stock_report
        data = dead_stock_report(self._get_tenant(request))
        return Response(data)

    @action(detail=False, methods=['get'], url_path='reports/inventory-aging')
    def inventory_aging(self, request):
        from inventory.services.dashboard_service import inventory_aging_report
        data = inventory_aging_report(self._get_tenant(request))
        return Response(data)

    @action(detail=False, methods=['get'], url_path='reports/top-suppliers')
    def top_suppliers(self, request):
        from inventory.services.dashboard_service import top_suppliers
        data = top_suppliers(self._get_tenant(request))
        return Response(data)

    @action(detail=False, methods=['get'], url_path='reports/most-adjusted')
    def most_adjusted(self, request):
        from inventory.services.dashboard_service import most_adjusted_items
        data = most_adjusted_items(self._get_tenant(request))
        return Response(data)

    @action(detail=False, methods=['get'], url_path='reports/most-transferred')
    def most_transferred(self, request):
        from inventory.services.dashboard_service import most_transferred_items
        data = most_transferred_items(self._get_tenant(request))
        return Response(data)

    # ---- Export ----

    @action(detail=False, methods=['get'], url_path='export/(?P<report_type>[^/.]+)')
    def export(self, request, report_type=None):
        """Export any report as CSV or Excel."""
        tenant_id = self._get_tenant(request)
        fmt = request.query_params.get('export_format', 'csv').lower()
        from inventory.services.dashboard_service import (
            current_stock_report, stock_ledger_report,
            stock_movement_report, stock_summary_report,
            inventory_valuation_report, reserved_stock_report,
            damaged_stock_report, adjustment_report,
            transfer_report, reservation_report,
            stock_count_report, purchase_order_report,
            goods_receipt_report, purchase_return_report,
            low_stock_report, out_of_stock_report,
            reorder_report, fast_moving_items,
            slow_moving_items, dead_stock_report,
            inventory_aging_report,
        )
        report_funcs = {
            'current-stock': current_stock_report,
            'stock-ledger': stock_ledger_report,
            'stock-movement': stock_movement_report,
            'stock-summary': stock_summary_report,
            'valuation': inventory_valuation_report,
            'reserved-stock': reserved_stock_report,
            'damaged-stock': damaged_stock_report,
            'adjustments': adjustment_report,
            'transfers': transfer_report,
            'reservations': reservation_report,
            'stock-counts': stock_count_report,
            'purchase-orders': purchase_order_report,
            'goods-receipts': goods_receipt_report,
            'purchase-returns': purchase_return_report,
            'low-stock': lambda t, f=None: low_stock_report(t),
            'out-of-stock': lambda t, f=None: out_of_stock_report(t),
            'reorder': lambda t, f=None: reorder_report(t),
            'fast-moving': lambda t, f=None: fast_moving_items(t),
            'slow-moving': lambda t, f=None: slow_moving_items(t),
            'dead-stock': lambda t, f=None: dead_stock_report(t),
            'inventory-aging': lambda t, f=None: inventory_aging_report(t),
        }
        func = report_funcs.get(report_type)
        if not func:
            return Response({'error': f'Unknown report type: {report_type}'},
                            status=status.HTTP_404_NOT_FOUND)
        data = func(tenant_id, self._get_filters(request))

        import csv, io as py_io
        if fmt == 'csv':
            output = py_io.StringIO()
            writer = csv.writer(output)
            if data and isinstance(data, list) and len(data) > 0:
                writer.writerow(data[0].keys())
                for row in data:
                    writer.writerow(row.values())
            output.seek(0)
            resp = HttpResponse(output.getvalue(), content_type='text/csv')
            resp['Content-Disposition'] = f'attachment; filename="{report_type}.csv"'
            return resp
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = report_type
            if data and isinstance(data, list) and len(data) > 0:
                ws.append(list(data[0].keys()))
                for row in data:
                    ws.append([str(v) if not isinstance(v, (int, float, str)) else v for v in row.values()])
            output = py_io.BytesIO()
            wb.save(output)
            output.seek(0)
            resp = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            resp['Content-Disposition'] = f'attachment; filename="{report_type}.xlsx"'
            return resp
